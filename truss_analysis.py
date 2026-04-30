
import openpyxl
import numpy as np
import pandas as pd

PJ = 0
JP = 0
COORD = 0
MSUP = 0
MPRP = 0
EM = 0
CP = 0
NCJT = 2  # number of coordinates per joint (2 for trusses)
grade = "SXXX"
shapecomp = "XXX"
shapeten = "XXX"
jointing = "XXXX"
stag = False
nh = 0.0
d = 0.0
s = 0.0
p = 0.0
ngs = 0.0
condition = ""
endcondition = ""


def inputxl(joints_file, members_file):
    global PJ, JP, COORD, MSUP, MPRP, EM, CP

    if joints_file is None or members_file is None:
        raise ValueError("Missing joints or members Excel file.")

    try:
        joints_wb = openpyxl.load_workbook(joints_file)
        members_wb = openpyxl.load_workbook(members_file)
    except Exception as e:
        raise ValueError(f"Error loading Excel files: {e}")

    joints = joints_wb.active
    members = members_wb.active

    memberno = int(members.cell(row=4, column=21).value)
    jointno = int(joints.cell(row=2, column=21).value)

    coord = [
        [float(joints.cell(row=i, column=2).value),
         float(joints.cell(row=i, column=3).value)]
        for i in range(2, jointno + 2)
    ]
    COORD = np.array(coord)

    msup = []
    for i in range(2, jointno + 2):
        if float(joints.cell(row=i, column=4).value) > 0.0:
            msup.append([
                float(joints.cell(row=i, column=1).value),
                float(joints.cell(row=i, column=9).value),
                float(joints.cell(row=i, column=10).value)
            ])
    MSUP = np.array(msup)

    mprp = [
        [float(members.cell(row=i, column=c).value) for c in range(2, 6)]
        for i in range(2, memberno + 2)
    ]
    MPRP = np.array(mprp)

    NMP = int(members.cell(row=2, column=21).value)
    EM = np.array([[float(members.cell(row=i, column=8).value)]
                   for i in range(2, NMP + 2)])

    NCP = int(members.cell(row=3, column=21).value)
    CP = np.array([[float(members.cell(row=i, column=9).value)]
                   for i in range(2, NCP + 2)])

    jp, pj = [], []
    for i in range(2, jointno + 2):
        if float(joints.cell(row=i, column=6).value) > 0.0:
            jp.append([float(joints.cell(row=i, column=1).value)])
            pj.append([
                float(joints.cell(row=i, column=7).value),
                float(joints.cell(row=i, column=8).value)
            ])

    JP = np.array(jp)
    PJ = np.array(pj)

def set_steel_properties(
    grade_in,
    shapeten_in,
    shapecomp_in,
    jointing_in,
    nh_in=0,
    d_in=0,
    staggered=False,
    s_in=0,
    p_in=0,
    ngs_in=0
):
    global grade, shapeten, shapecomp, jointing
    global stag, nh, d, s, p, ngs

    grade = grade_in
    shapeten = shapeten_in
    shapecomp = shapecomp_in
    jointing = jointing_in

    if jointing == "bolt":
        stag = staggered
        nh = nh_in
        d = d_in

        if stag:
            s = s_in
            p = p_in
            ngs = ngs_in
        else:
            s, p, ngs = 0, 0, 0

    elif jointing == "weld":
        stag = False
        nh, d, s, p, ngs = 0, 0, 0, 0, 0

    else:
        raise ValueError("Jointing must be 'bolt' or 'weld'")


def table_reader(table, val):
    if table == "CHS":
        ex = openpyxl.load_workbook("CHS.xlsx")
        ex = ex.active
        n = 0.0
        i = 1
        while n<val:
            i = i + 1
            n = float(ex.cell(row=i, column=4).value)
        dia = float(ex.cell(row=i, column=1).value)
        thickness = float(ex.cell(row=i, column=2).value)
        area = n
        rg = float(ex.cell(row=i, column=7).value)*10.0
        return area, dia, thickness, rg
    elif table == "Angle":
        ex = openpyxl.load_workbook("Angle.xlsx")
        ex = ex.active
        n = 0.0
        i = 44
        while n<=val:
            i = i - 1
            n = float(ex.cell(row=i, column=6).value)
        size = ex.cell(row=i, column=1).value
        thickness = float(ex.cell(row=i, column=2).value)
        area = n
        return area, size, thickness
    elif table == "UC":
        ex = openpyxl.load_workbook("UC-2.xlsx").active
        n = 0.0
        i = 1
        while n<val:
            i = i + 1
            n = float(ex.cell(row=i, column=14).value)
        area = n
        size = ex.cell(row=i, column=1).value
        Iy = float(ex.cell(row=i, column=2).value)*10000.0
        Iz = float(ex.cell(row=i, column=3).value)*10000.0
        ry = float(ex.cell(row=i, column=4).value)*10.0
        rz = float(ex.cell(row=i, column=5).value)*10.0
        if Iy<Iz:
            I = Iy
            R = ry
            axis = "y"
        else:
            axis = "z"
            I = Iz
            R = rz
        return area, size, I, R, axis
    elif table == "UB":
        ex = openpyxl.load_workbook("UB-2.xlsx").active
        n = 0.0
        i = 1
        while n < val:
            i = i + 1
            n = float(ex.cell(row=i, column=14).value)
        area = n
        size = ex.cell(row=i, column=1).value
        Iy = float(ex.cell(row=i, column=2).value) * 10000.0
        Iz = float(ex.cell(row=i, column=3).value) * 10000.0
        ry = float(ex.cell(row=i, column=4).value) * 10.0
        rz = float(ex.cell(row=i, column=5).value) * 10.0
        if Iy < Iz:
            I = Iy
            R = ry
            axis = "y"
        else:
            axis = "z"
            I = Iz
            R = rz
        return area, size, I, R, axis

    else:
        print("\n\nTable not recognized\n\n")
        return None

def ten_designer(f):
    global grade, shapeten, jointing, stag, nh, d, s, p, ngs
    grades = openpyxl.load_workbook("grades.xlsx").active
    i = 1
    pop = "sjbdcusdd"
    while pop!=grade and i<=7:
        i = i+1
        pop = grades.cell(row=i, column=1).value
    if i>7:
        print("\n\nGrade not recognized\n\n")
        return None
    fy = float(grades.cell(row=i, column=2).value)
    fu = float(grades.cell(row=i, column=3).value)
    areq = f/fy
    anet = 0.0
    agross = 0.0
    counter = 0
    while anet<areq and counter<=1000:
        counter = counter + 1
        section = table_reader(shapeten, max(agross,areq))
        if section is None:
            return None
        agross = section[0]
        t = section[2]
        if jointing == "weld":
            anet = agross
        elif jointing == "bolt" and not stag:
            anet = agross - nh * t * (d + 2.0)
        elif jointing == "bolt" and stag:
            anet = agross - nh * t * (d + 2.0) + ngs * (((s ** 2) * t) / (4 * p))
        else:
            print("\n\nJointing not recognized.\n\n")
    if counter==1000:
        print("\n\nInfinite loop in section sizing. - 1\n\n")
        return None
    npl = (agross*fy)/1.0
    nu = (0.9*anet*fu)/1.25
    counter = 1
    while f>min(npl,nu) and counter<=1000:
        counter = counter + 1
        section = table_reader(shapeten, agross)
        if section is None:
            return None
        agross = section[0]
        t = section[2]
        if jointing == "weld":
            anet = agross
        elif jointing == "bolt" and not stag:
            anet = agross - nh * t * (d + 2.0)
        elif jointing == "bolt" and stag:
            anet = agross - nh * t * (d + 2.0) + ngs * (((s ** 2) * t) / (4 * p))
        else:
            print("\n\nSomething went wrong. Error code 1\n\n")
            return None
        npl = (agross * fy) / 1.0
        nu = (0.9 * anet * fu) / 1.25
    if counter==1000:
        print("\n\nInfinite loop in section sizing. - 2\n\n")
        return None
    return section


def comp_designer(f,L):
    global grade, shapecomp, stag, nh, d, s, p, ngs
    grades = openpyxl.load_workbook("grades.xlsx").active
    i = 1
    pop = "sjbdcusdd"
    while pop != grade and i <= 7:
        i = i + 1
        pop = grades.cell(row=i, column=1).value
    if i > 7:
        print("Grade not recognized")
        return None
    fy = float(grades.cell(row=i, column=2).value)
    E = 210000  # MPa
    if shapecomp == "UC" or shapecomp == "UB":
        alpha = 0.34
    elif shapecomp == "CHS":
        alpha = 0.21
    elif shapecomp == "Angle":
        alpha = 0.49
    else:
        print("\n\nCompression shape not recognized\n\n")
        return None
    if shapecomp == "UB" or shapecomp == "UC":
        agross = f/fy
        counter = 0
        while counter <= 1000:
            counter += 1
            section = table_reader(shapecomp, agross)
            if section is None:
                return None
            A, size, I, r, axis = section
            Ncr = (np.pi ** 2 * E * I) / (L ** 2)
            lam_bar = np.sqrt((A * fy) / Ncr)
            phi = 0.5 * (1 + alpha * (lam_bar - 0.2) + lam_bar ** 2)
            chi = min(1 / (phi + np.sqrt(phi ** 2 - lam_bar ** 2)), 1.0)
            NbRd = chi * A * fy
            if NbRd >= f:
                return {
                    "area": A,
                    "size": size,
                    "I": I,
                    "r": r,
                    "axis": axis,
                    "lambda_bar": lam_bar,
                    "chi": chi,
                    "NbRd": NbRd
                }
            else:
                agross = A
    elif shapecomp == "CHS":

        agross = f / fy
        counter = 0

        while counter <= 1000:
            counter += 1
            section = table_reader(shapecomp, agross)
            if section is None:
                return None
            A, dia, t, r = section

            # For CHS: I = A * r^2
            I = A * (r ** 2)
            Ncr = (np.pi ** 2 * E * I) / (L ** 2)
            lam_bar = np.sqrt((A * fy) / Ncr)
            phi = 0.5 * (1 + alpha * (lam_bar - 0.2) + lam_bar ** 2)
            chi = min(1 / (phi + np.sqrt(max(phi ** 2 - lam_bar ** 2, 0))), 1.0)
            NbRd = chi * A * fy
            if NbRd >= f:
                return {
                    "area": A,
                    "size": f"CHS {dia}x{t}",
                    "I": I,
                    "r": r,
                    "axis": "circular",
                    "lambda_bar": lam_bar,
                    "chi": chi,
                    "NbRd": NbRd
                }

            else:
                agross = A












def assign_structure_coordinates():
    global COORD, MSUP
    NJ = COORD.shape[0]
    NR = int(np.sum(MSUP[:, 1:]))
    NDOF = NCJT * NJ - NR
    NSC = np.zeros(NCJT * NJ, dtype=int)
    J = 0
    K = NDOF

    for i in range(NJ):
        is_support = np.where(MSUP[:, 0].astype(int) == i + 1)[0]
        if len(is_support) == 0:
            for j in range(NCJT):
                J += 1
                NSC[i * NCJT + j] = J
        else:
            sup_idx = is_support[0]
            for j in range(NCJT):
                if int(MSUP[sup_idx, j + 1]) == 1:
                    K += 1
                    NSC[i * NCJT + j] = K
                else:
                    J += 1
                    NSC[i * NCJT + j] = J
    return NSC, NDOF

def generate_stiffness_matrix(NSC, NDOF):
    global COORD, MPRP, EM, CP
    S = np.zeros((NDOF, NDOF))
    for m in range(MPRP.shape[0]):
        jb, je, mat_idx, cp_idx = map(lambda x: int(x) - 1, MPRP[m])
        E = EM[int(mat_idx), 0]
        A = CP[int(cp_idx), 0]
        x1, y1 = COORD[jb]
        x2, y2 = COORD[je]
        L = np.hypot(x2 - x1, y2 - y1)
        CX = (x2 - x1) / L
        CY = (y2 - y1) / L
        z = E * A / L
        c2, s2, cs = CX ** 2, CY ** 2, CX * CY
        k = z * np.array([[ c2, cs, -c2, -cs],
                          [ cs, s2, -cs, -s2],
                          [-c2, -cs, c2, cs],
                          [-cs, -s2, cs, s2]])
        idx = [NSC[jb * NCJT] - 1, NSC[jb * NCJT + 1] - 1,
               NSC[je * NCJT] - 1, NSC[je * NCJT + 1] - 1]
        for i in range(4):
            for j in range(4):
                if idx[i] < NDOF and idx[j] < NDOF:
                    S[idx[i], idx[j]] += k[i, j]
    return S

def calculate_member_lengths():
    global COORD, MPRP

    member_lengths = {}

    for m in range(MPRP.shape[0]):
        member_id = m + 1
        jb, je = int(MPRP[m,0]) - 1, int(MPRP[m,1]) - 1

        x1, y1 = COORD[jb]
        x2, y2 = COORD[je]

        L = np.hypot(x2 - x1, y2 - y1)

        member_lengths[member_id] = L

    return member_lengths

def form_load_vector(NSC, NDOF):
    global JP, PJ
    P = np.zeros(NDOF)
    for i in range(JP.shape[0]):
        joint = int(JP[i, 0]) - 1
        fx, fy = PJ[i]
        idx_x = NSC[joint * NCJT] - 1
        idx_y = NSC[joint * NCJT + 1] - 1
        if idx_x < NDOF: P[idx_x] = fx
        if idx_y < NDOF: P[idx_y] = fy
    return P

def solve_displacements(S, P):
    return np.linalg.solve(S, P)

def calculate_member_forces(NSC, D):
    global COORD, MPRP, EM, CP
    member_forces = {}

    for m in range(MPRP.shape[0]):
        member_id = m + 1
        jb, je, mat_idx, cp_idx = map(lambda x: int(x) - 1, MPRP[m])
        E = EM[int(mat_idx), 0]
        A = CP[int(cp_idx), 0]
        x1, y1 = COORD[jb]
        x2, y2 = COORD[je]
        L = np.hypot(x2 - x1, y2 - y1)
        CX = (x2 - x1) / L
        CY = (y2 - y1) / L

        idx = [NSC[jb * NCJT] - 1, NSC[jb * NCJT + 1] - 1,
               NSC[je * NCJT] - 1, NSC[je * NCJT + 1] - 1]

        u = np.array([D[i] if i < len(D) else 0.0 for i in idx])
        delta = CX * (u[2] - u[0]) + CY * (u[3] - u[1])
        axial_force = E * A / L * delta

        member_forces[member_id] = axial_force

    return member_forces


#def run_truss_analysis(verbose=False):
    inputxl()
    NSC, NDOF = assign_structure_coordinates()
    S = generate_stiffness_matrix(NSC, NDOF)
    P = form_load_vector(NSC, NDOF)
    D = solve_displacements(S, P)

    member_forces = calculate_member_forces(NSC, D)

    # Always show clean axial force table
    member_ids = [f"Member {m}" for m in member_forces.keys()]
    force_values = list(member_forces.values())

    force_table = pd.DataFrame({
        "Member": member_ids,
        "Axial Force": force_values
    })

    print("\nMember Forces:")
    print(force_table.to_string(index=False))

    # Only show detailed design output if requested
    if verbose:
        print("\nDetailed Design Results:")
        for m, res in design_results.items():
            print(f"Member {m}: {res}")

    return member_forces, design_results#

def design_single_member(force, member_type, length=1000):
    """
    Designs a single steel member.

    force: kN
    member_type: "Tension" or "Compression"
    length: mm (only needed for compression)
    """

    f = abs(force) * 1000  # convert to N

    if member_type == "Tension":

        section = ten_designer(f)

        if section is None:
            return None

        area = section[0]
        size = section[1]
        thickness = section[2]

        return {
            "Type": "Tension",
            "Shape": shapeten,
            "Size": size,
            "Thickness": thickness,
            "Area": area
        }

    elif member_type == "Compression":

        section = comp_designer(f, length)

        if section is None:
            return None

        return {
            "Type": "Compression",
            "Shape": shapecomp,
            "Size": section["size"],
            "Axis": section["axis"],
            "χ": round(section["chi"], 3),
            "NbRd (kN)": round(section["NbRd"]/1000, 3),
            "Capacity Utilization (%)": round((force/(section["NbRd"]/1000))*100, 3)
        }

    else:
        return None

def run_analysis_and_design_table(joints_file, members_file):

    if joints_file is None or members_file is None:
        raise ValueError("All three Excel files must be uploaded.")

    # ---- Structural Analysis ----
    inputxl(joints_file, members_file)

    NSC, NDOF = assign_structure_coordinates()

    S = generate_stiffness_matrix(NSC, NDOF)

    P = form_load_vector(NSC, NDOF)

    D = solve_displacements(S, P)

    member_forces = calculate_member_forces(NSC, D)

    # ---- Design containers ----
    tension_results = []
    compression_results = []

    member_lengths = calculate_member_lengths()

    for member_id, force in member_forces.items():

        L = member_lengths[member_id]

        if force > 0:

            section = ten_designer(force * 1000)

            if section is not None:

                area = section[0]
                size = section[1]
                thickness = section[2]

                tension_results.append({
                    "Member": f"Member {member_id}",
                    "Force (kN)": round(force, 3),
                    "Shape": shapeten,
                    "Size": size,
                    "Thickness": thickness
                })

        elif force < 0:

            section = comp_designer(-force * 1000, L)

            if section is not None:

                compression_results.append({
                    "Member": f"Member {member_id}",
                    "Force (kN)": round(force, 3),
                    "Shape": shapecomp,
                    "Size": section["size"],
                    "Axis": section["axis"],
                    "Capacity Utilization (%)": round((-force/(section["NbRd"]/1000))*100, 3),
                    "χ": round(section["chi"], 3),
                    "NbRd (kN)": round(section["NbRd"]/1000, 3)
                })

    tension_table = pd.DataFrame(tension_results)
    compression_table = pd.DataFrame(compression_results)

    return tension_table, compression_table

def get_beam_Iy(size):
    """Return Iy (mm⁴) for the given UB section designation, or None if not found."""
    try:
        ex = openpyxl.load_workbook("UB-2.xlsx").active
        for i in range(2, ex.max_row + 1):
            if ex.cell(row=i, column=1).value == size:
                return float(ex.cell(row=i, column=2).value) * 10000  # cm⁴ → mm⁴
    except Exception:
        pass
    return None


def restrained_beam(M, V, L=None):
    """
    Design a restrained beam to EC3.

    M  – design bending moment (kNm)
    V  – design shear force (kN)
    L  – beam span (mm), optional.  When provided the SLS deflection check
         δ ≤ L/300 is performed and enforced; sections that pass strength but
         fail deflection are skipped so that the returned section satisfies both
         ULS and SLS requirements.
    """
    global grade

    grades = openpyxl.load_workbook("grades.xlsx").active

    # ---- Get fy ----
    for i in range(2, 10):
        if grades.cell(row=i, column=1).value == grade:
            fy = float(grades.cell(row=i, column=2).value)
            break

    gamma_M0 = 1.0
    E_steel = 210000.0  # N/mm²

    ex = openpyxl.load_workbook("UB-2.xlsx").active

    i = 1

    while True:
        i += 1

        size = ex.cell(row=i, column=1).value

        if size is None:
            raise ValueError("No suitable section found in UB table.")

        Wyy = float(ex.cell(row=i, column=8).value) * 1000  # mm³
        Wzz = float(ex.cell(row=i, column=9).value) * 1000
        if Wyy > Wzz:
            axis = "y"
        else:
            axis = "z"
        Wpl = max(Wyy, Wzz)
        Iy  = float(ex.cell(row=i, column=2).value) * 10000  # cm⁴ → mm⁴
        tw = float(ex.cell(row=i, column=15).value)           # web thickness
        h = float(ex.cell(row=i, column=16).value)            # depth
        tf = float(ex.cell(row=i, column=17).value)
        hw = h - 2 * tf
        Av = hw * tw  # mm²

        Aw = hw * tw

        # ---- SHEAR CHECK ----
        Vpl_Rd = (Av * fy) / (np.sqrt(3) * gamma_M0) / 1000  # kN

        if V > Vpl_Rd:
            continue  # FAIL

        # ---- MOMENT CAPACITY ----
        Mpl_Rd = (Wpl * fy) / gamma_M0 / 1e6  # kNm

        if V / Vpl_Rd <= 0.5:
            M_Rd = Mpl_Rd

        else:
            rho = (2 * V / Vpl_Rd - 1) ** 2

            M_Rd = (
                (Wpl - (rho * Aw**2) / (4 * tw)) * fy / gamma_M0
            ) / 1e6  # kNm

        # ---- ULS MOMENT CHECK ----
        if M > M_Rd:
            continue

        # ---- SLS DEFLECTION CHECK (EC3, δ ≤ L/300) ----
        # Estimate mid-span deflection for a simply-supported beam under
        # equivalent UDL: δ = 5ML²/(48EI)  [M in N·mm, L in mm, E in N/mm², I in mm⁴]
        if L is not None:
            M_Nmm  = M * 1e6           # kNm → N·mm
            delta  = 5.0 * M_Nmm * L ** 2 / (48.0 * E_steel * Iy)   # mm
            delta_lim = L / 300.0      # mm  (EC3 serviceability limit)
            if delta > delta_lim:
                continue               # fails deflection → try next (larger) section

            return {
                "Type": "Restrained Beam",
                "Size": size,
                "M_Rd (kNm)":         round(M_Rd, 2),
                "V_Rd (kN)":          round(Vpl_Rd, 2),
                "Utilization (%)":    round((M / M_Rd) * 100, 2),
                "delta (mm)":         round(delta, 2),
                "delta_lim (mm)":     round(delta_lim, 2),
                "Deflection Check":   "PASS ✓",
            }

        return {
            "Type": "Restrained Beam",
            "Size": size,
            "M_Rd (kNm)":      round(M_Rd, 2),
            "V_Rd (kN)":       round(Vpl_Rd, 2),
            "Utilization (%)": round((M / M_Rd) * 100, 2),
        }

def unrestrained_beam(M, V, L):
    global condition, endcondition, grade

    # ---- Load steel grade ----
    grades = openpyxl.load_workbook("grades.xlsx").active

    for i in range(2, 10):
        if grades.cell(row=i, column=1).value == grade:
            fy = float(grades.cell(row=i, column=2).value)
            break

    if 'fy' not in locals():
        raise ValueError("Steel grade not found in grades.xlsx")

    E = 210000  # MPa
    gamma_M1 = 1.0

    # ---- Load UB table ----
    ex = openpyxl.load_workbook("UB-2.xlsx").active

    i = 1

    while True:
        i += 1

        size = ex.cell(row=i, column=1).value

        # ---- Stop if no more sections ----
        if size is None:
            raise ValueError("No suitable section found in UB table.")

        # ---- Section properties ----
        Wyy = float(ex.cell(row=i, column=8).value) * 1000  # mm³
        Wpl = Wyy  # Use major axis only

        Iy = float(ex.cell(row=i, column=2).value) * 10000  # mm⁴
        Iz = float(ex.cell(row=i, column=3).value) * 10000
        I = Iy

        tw = float(ex.cell(row=i, column=15).value)
        h = float(ex.cell(row=i, column=16).value)
        tf = float(ex.cell(row=i, column=17).value)
        b = float(ex.cell(row=i, column=19).value)
        iz = float(ex.cell(row=i, column=5).value)*10.0
        if endcondition=="Free":
            k = 1.0
        elif endcondition=="Partial":
            k = 0.85
        elif endcondition=="Full":
            k = 0.7
        elif endcondition=="Cantilever":
            k = 2.0
        else:
            raise ValueError("Endcondition not recognized.")
        lamz = (k*L)/iz
        laml = np.pi * np.sqrt(E/fy)
        lamzba = lamz / laml
        c1 = 1.0
        u = 0.9
        vee = 1.0
        bew = 1.0
        lamltb = (1/np.sqrt(c1))*u*vee*lamzba*np.sqrt(bew)
        hoverb = h/b
        if condition == "Rolled":
            if hoverb <= 2:
                alt = 0.34
            else:
                alt = 0.49
            phi = 0.5 * (1 + alt*(lamltb - 0.4) + 0.75*(lamltb**2))
            chi = min(1 / (phi + np.sqrt(phi**2 - 0.75*(lamltb**2))),1.0)
        else:
            if hoverb <= 2 and condition == "Welded":
                alt = 0.49
            elif hoverb > 2 and condition == "Welded":
                alt = 0.76
            else:
                alt = 0.76
            phi = 0.5 * (1 + alt*(lamltb - 0.2) + (lamltb**2))
            chi = min(1 / (phi + np.sqrt(phi**2 - (lamltb**2))),1.0)
        Mbrd = chi * Wpl * (fy/gamma_M1)
        # NOTE:
        # Mbrd is computed twice in this function.
        # The first computation (above) is in N·mm and is overwritten later by a kNm version.
        # This is redundant and potentially confusing.
        # TODO: Remove the earlier Mbrd definition and keep only the final (post-shear reduction) calculation.

        hw = h - 2 * tf
        Av = hw * tw  # shear area (mm²)
        Aw = hw * tw  # web area for reduction

        # ---- SHEAR CHECK ----
        Vpl_Rd = (Av * fy) / (np.sqrt(3) * gamma_M1) / 1000  # kN (Note: V is in kN)

        if V > Vpl_Rd:
            continue  # reject section

        # ---- HIGH SHEAR REDUCTION ----
        # If V is more than 50% of Vpl_Rd, reduce the moment capacity
        if V > 0.5 * Vpl_Rd:
            rho = ((2 * V / Vpl_Rd) - 1) ** 2
            # Reduced plastic modulus (Wpl_reduced)
            # Subtract the 'lost' capacity of the web area
            # Formula: M_y,V,Rd = (Wpl - (rho * Aw^2 / 4tw)) * fy / gamma_M0
            W_reduced = Wpl - (rho * (Aw ** 2)) / (4 * tw)
        else:
            W_reduced = Wpl

        # ---- UPDATED DESIGN MOMENT ----
        # Use W_reduced instead of Wpl in your Mbrd calculation
        # Note: Divide by 10^6 if W is mm3 and fy is MPa to get kNm
        Mbrd = (chi * W_reduced * fy / gamma_M1) / 1e6

        # ---- ULS MOMENT CHECK ----
        if M > Mbrd:
            continue

        # ---- SLS DEFLECTION CHECK (EC3, δ ≤ L/300) ----
        # Mid-span deflection estimate under equivalent UDL:
        # δ = 5ML²/(48EI)  [M in N·mm, L in mm, E in N/mm², I in mm⁴]
        E_steel = 210000.0  # N/mm²
        M_Nmm   = M * 1e6           # kNm → N·mm
        delta   = 5.0 * M_Nmm * L ** 2 / (48.0 * E_steel * Iy)   # mm
        delta_lim = L / 300.0       # mm  (EC3 serviceability limit)
        if delta > delta_lim:
            continue                # fails deflection → try next (larger) section

        return {
            "Type": "Unrestrained Beam",
            "Size": size,
            "x_LT": round(chi, 3),
            "Mb_Rd (kNm)":       round(Mbrd, 2),
            "Vpl_Rd":            round(Vpl_Rd, 2),
            "Utilization (%)":   round((M / Mbrd) * 100, 2),
            "delta (mm)":        round(delta, 2),
            "delta_lim (mm)":    round(delta_lim, 2),
            "Deflection Check":  "PASS ✓",
        }

        # # ---- SHEAR REDUCTION ----
        # if V / Vpl_Rd <= 0.5:
        #      W_eff = Wpl
        # else:
        #      rho = (2 * V / Vpl_Rd - 1) ** 2
        #      W_eff = max(Wpl - (rho * Aw**2) / (4 * tw), 0.1 * Wpl)
        #
        # # ---- LTB (improved approximation) ----
        # Mcr = 2.5 * (np.pi**2 * E * I) / (L**2)
        #
        # lam = np.sqrt((W_eff * fy) / Mcr)
        #
        # alpha = 0.34
        # phi = 0.5 * (1 + alpha * (lam - 0.2) + lam**2)
        #
        # chi = 1 / (phi + np.sqrt(max(phi**2 - lam**2, 0)))
        #
        # # ---- DESIGN MOMENT ----
        # Mb_Rd = chi * W_eff * fy / gamma_M1 / 1e6  # kNm
        #
        # # ---- CHECK ----
        # if M <= Mb_Rd:
        #     return {
        #         "Type": "Unrestrained Beam",
        #         "Size": size,
        #         "χ_LT": round(chi, 3),
        #         "Mb_Rd (kNm)": round(Mb_Rd, 2),
        #         "Utilization (%)": round((M / Mb_Rd) * 100, 2)
        #     }



def section_class(fy, section, Ned):
    """
    Classifies an I/H section (UC/UB) to EC3 (Class 1–4)

    Parameters:
    H   : overall depth (mm)
    b   : flange width (mm)
    tf  : flange thickness (mm)
    tw  : web thickness (mm)
    fy  : yield strength (MPa)
    Ned : axial force (N) (compression +ve)
    A   : area (mm²)

    Returns:
    int → section class (1 to 4)
    """

    # -------------------------
    # BASIC PARAMETERS
    # -------------------------
    H = section.h
    b = section.b
    tf = section.tf
    tw = section.tw
    A = section.A
    eps = np.sqrt(235.0 / fy)

    d = H - 2.0 * tf                 # clear web depth
    c = (b - tw) / 2.0              # flange outstand

    # Avoid division issues
    if A * fy == 0:
        return 4

    alpha = 0.5 * (1.0 + Ned / (A * fy))

    # Clamp alpha to sensible EC3 range
    alpha = np.clip(alpha, 0.0, 1.0)

    # -------------------------
    # FLANGE CLASS
    # -------------------------
    lambda_f = c / tf

    if lambda_f <= 9.0 * eps:
        flange_class = 1
    elif lambda_f <= 10.0 * eps:
        flange_class = 2
    elif lambda_f <= 14.0 * eps:
        flange_class = 3
    else:
        flange_class = 4

    # -------------------------
    # WEB CLASS
    # -------------------------
    lambda_w = d / tw

    if alpha > 0.5:
        denom = (13.0 * alpha - 1.0)

        # Prevent divide-by-zero or negative weirdness
        if denom <= 0:
            return 4

        limit1 = 396.0 * eps / denom
        limit2 = 456.0 * eps / denom
        # EC3 Table 5.2 — Class 3 web limit for the compression-dominant range (α > 0.5).
        # ψ_EC3 is the stress ratio at the web faces (tension-positive convention):
        #   ψ_EC3 = 1 − 1/α  maps α ∈ (0.5, 1.0) → ψ_EC3 ∈ (−1, 0)
        # The denominator 0.67 + 0.33·ψ_EC3 = 1 − 0.33/α is bounded in [0.34, 0.67]
        # for α ∈ (0.5, 1), so division by zero cannot occur.
        # α = 1 (full compression, ψ_EC3 = 1) is treated separately as 42ε.
        if alpha >= 1.0:
            limit3 = 42.0 * eps                                 # EC3: ψ=1, uniform compression
        else:
            # 0.5 < alpha < 1.0 — neutral axis is within the web
            psi_EC3 = 1.0 - 1.0 / alpha                        # ψ_EC3 ∈ (−1, 0)
            ec3_limit3 = 42.0 * eps / (0.67 + 0.33 * psi_EC3)  # 124ε at α→0.5, 62.7ε at α→1
            limit3 = max(limit2, ec3_limit3)                    # ensure Class 3 ≥ Class 2

    else:
        # Prevent division by zero
        if alpha <= 0:
            return 4

        limit1 = 36.0 * eps / alpha
        limit2 = 41.5 * eps / alpha
        # EC3 Table 5.2 — Class 3 web limit for tension-dominant case (α ≤ 0.5).
        # At α = 0.5 (pure bending, ψ_EC3 = −1): EC3 gives 124ε.
        # For α < 0.5 the compression zone is smaller, so the limit is even more
        # permissive; 124ε is a conservative (safe) lower bound.
        limit3 = max(limit2, 124.0 * eps)

    if lambda_w <= limit1:
        web_class = 1
    elif lambda_w <= limit2:
        web_class = 2
    elif lambda_w <= limit3:
        web_class = 3
    else:
        web_class = 4

    # -------------------------
    # FINAL CLASS
    # -------------------------
    return int(max(flange_class, web_class))

def eff_L(L, endcondition):
    if endcondition=="Pinned-Pinned":
        return L
    elif endcondition=="Fixed-Fixed":
        return L*0.5
    elif endcondition=="Fixed-Pinned":
        return L*0.7
    elif endcondition=="Fixed-Free":
        return L*2.0
    else:
        return None

class Section:
    def __init__(self, designation, A, Iy, Iz, Wpl, Wel, Wpl_z, Wel_z, Iw, It, tw, h, tf, b, iz, seclass=None, alphay=None, alphaz=None):
        self.designation = designation
        self.A = A
        self.Iy = Iy
        self.Iz = Iz
        self.Wpl = Wpl       # major-axis plastic modulus W_pl,y
        self.Wel = Wel       # major-axis elastic modulus W_el,y
        self.Wpl_z = Wpl_z  # minor-axis plastic modulus W_pl,z
        self.Wel_z = Wel_z  # minor-axis elastic modulus W_el,z
        self.Iw = Iw
        self.It = It
        self.tw = tw
        self.h = h
        self.tf = tf
        self.b = b
        self.iz = iz
        self.seclass = seclass
        self.alphay = alphay
        self.alphaz = alphaz




def beam_column_table(shape, i):
    global condition, grade
    if shape == "UB":
        ex = openpyxl.load_workbook("UB-2.xlsx").active
        i = max(i, 17)
    elif shape == "UC":
        ex = openpyxl.load_workbook("UC-2.xlsx").active
    designation = ex.cell(row=i, column=1).value
    A = float(ex.cell(row=i, column=14).value)
    Iy = float(ex.cell(row=i, column=2).value) * 10000
    Iz = float(ex.cell(row=i, column=3).value) * 10000
    Wpl = float(ex.cell(row=i, column=8).value) * 1000    # W_pl,y  (major axis, cm³ → mm³)
    Wel = float(ex.cell(row=i, column=6).value) * 1000    # W_el,y  (major axis, cm³ → mm³)
    Wpl_z = float(ex.cell(row=i, column=9).value) * 1000  # W_pl,z  (minor axis, cm³ → mm³)
    Wel_z = float(ex.cell(row=i, column=7).value) * 1000  # W_el,z  (minor axis, cm³ → mm³)
    Iw = float(ex.cell(row=i, column=12).value) * (100 ** 6)
    It = float(ex.cell(row=i, column=13).value) * 10000
    tw = float(ex.cell(row=i, column=15).value)
    h = float(ex.cell(row=i, column=16).value)
    tf = float(ex.cell(row=i, column=17).value)
    b = float(ex.cell(row=i, column=19).value)
    iz = float(ex.cell(row=i, column=5).value) * 10.0
    if condition == "Rolled":
        if (h/b)<=1.2 and not grade=="S460":
            if tf<=100:
                alphay = 0.34
                alphaz = 0.49
            else:
                alphay = 0.76
                alphaz = 0.76
        elif (h/b)>1.2 and not grade=="S460":
            if tf<=40:
                alphay = 0.21
                alphaz = 0.34
            else:
                alphay = 0.34
                alphaz = 0.49
        else:
            if tf<=40:
                alphay = 0.13
                alphaz = 0.13
            elif tf>100:
                alphay = 0.49
                alphaz = 0.49
            else:
                alphay = 0.21
                alphaz = 0.21
    else:
        if tf<=40:
            alphay = 0.34
            alphaz = 0.49
        else:
            alphay = 0.49
            alphaz = 0.76



    outer = Section(designation, A, Iy, Iz, Wpl, Wel, Wpl_z, Wel_z, Iw, It, tw, h, tf, b, iz)
    outer.alphay = alphay
    outer.alphaz = alphaz
    return outer


def beam_column(L, Ned, Mzed, Myed, shape, C1, all_axis_similar=True):
    global condition, grade
    grades = openpyxl.load_workbook("grades.xlsx").active

    for i in range(2, 10):
        if grades.cell(row=i, column=1).value == grade:
            fy = float(grades.cell(row=i, column=2).value)
            break

    if 'fy' not in locals():
        raise ValueError("Steel grade not found in grades.xlsx")
    if all_axis_similar:
        Lcry = eff_L(L, endcondition)
        Lcrz = Lcry
    else:
        Lcrz = eff_L(L, endcondition[0])
        Lcry = eff_L(L, endcondition[1])
    E = 210000
    G = 81000
    i = 1
    if shape == "UB":
        ex = openpyxl.load_workbook("UB-2.xlsx").active
    elif shape == "UC":
        ex = openpyxl.load_workbook("UC-2.xlsx").active
    else:
        raise ValueError("Unknown shape")
    A0 = Ned / fy
    A = 0.0
    while True:
        while not A > A0:
            i += 1
            A = float(ex.cell(row=i, column=14).value)
        pop = beam_column_table(shape, i)
        pop.seclass = section_class(fy, pop, Ned)
        if pop.seclass == 1 or pop.seclass == 2:
            Nrd = (pop.A * fy)
            Mrd = pop.Wpl * fy
        elif pop.seclass == 3:
            Nrd = (pop.A * fy)
            Mrd = pop.Wel * fy
        elif pop.seclass == 4:
            i += 1
            continue
        Ncry = ((np.pi ** 2) * E * pop.Iy) / (Lcry ** 2)
        Ncrz = ((np.pi ** 2) * E * pop.Iz) / (Lcrz ** 2)
        lamy = np.sqrt((pop.A * fy) / Ncry)
        lamz = np.sqrt((pop.A * fy) / Ncrz)
        phiy = 0.5 * (1 + pop.alphay * (lamy - 0.2) + (lamy ** 2))
        phiz = 0.5 * (1 + pop.alphaz * (lamz - 0.2) + (lamz ** 2))
        chiy = min(1 / (phiy + np.sqrt(phiy ** 2 - lamy ** 2)), 1.0)
        chiz = min(1 / (phiz + np.sqrt(phiz ** 2 - lamz ** 2)), 1.0)
        term2 = (pop.A * fy)
        chi = min(chiy, chiz)
        Nbrd = chi * term2
        # -------------------------
        # LATERAL TORSIONAL BUCKLING
        # -------------------------
        LcrLT = Lcry  # assumption (you can refine later)

        Mcr = C1 * ((np.pi ** 2 * E * pop.Iz) / (LcrLT ** 2)) * np.sqrt(
            (pop.Iw / pop.Iz) +
            ((LcrLT ** 2 * G * pop.It) / (np.pi ** 2 * E * pop.Iz))
        )

        # Select correct major-axis section modulus (EC3 Cl. 6.3.2.1)
        if pop.seclass in [1, 2]:
            Wy = pop.Wpl
        else:
            Wy = pop.Wel

        lamLT = np.sqrt((Wy * fy) / Mcr)

        # LTB imperfection factor per EC3 Table 6.3 (buckling curve b → α_LT = 0.34)
        alpha_LT = 0.34

        phi_LT = 0.5 * (1 + alpha_LT * (lamLT - 0.2) + lamLT ** 2)
        chi_LT = min(1 / (phi_LT + np.sqrt(phi_LT ** 2 - lamLT ** 2)), 1.0)

        Mbrd = chi_LT * Wy * fy

        # -------------------------
        # MINOR AXIS BENDING RESISTANCE (EC3 Cl. 6.2.5)
        # Use the correct minor-axis section modulus W_pl,z / W_el,z
        # -------------------------
        if pop.seclass in [1, 2]:
            Wz = pop.Wpl_z  # minor-axis plastic modulus
        else:
            Wz = pop.Wel_z  # minor-axis elastic modulus

        Mzrd = Wz * fy

        # ===================== MODIFIED INTERACTION FACTORS =====================
        # Calculate interaction factors according to Eurocode 3 Annex B

        # 1. Calculate C_m factors (moment distribution factors)
        # Assuming uniform moment distribution (most conservative)
        # For more accurate values, you'd need moment diagram information
        psi_y = 1.0  # Ratio of end moments (1.0 = uniform moment)
        psi_z = 1.0

        C_my = max(0.4, 0.6 + 0.4 * psi_y)
        C_mz = max(0.4, 0.6 + 0.4 * psi_z)
        C_mz = C1
        C_mLT = max(0.4, 0.6 + 0.4 * psi_z)
        C_mLT = C1

        # 2. Calculate normalized axial force
        # Note: Using N_Rk = Nrd (plastic resistance)
        N_Rk = Nrd
        gamma_M1 = 1.0  # Partial safety factor (use appropriate value from material specs)

        # Calculate n_y and n_z using the appropriate reduction factors
        n_y = Ned / (chiy * N_Rk / gamma_M1) if chiy > 0 else 0
        n_z = Ned / (chiz * N_Rk / gamma_M1) if chiz > 0 else 0

        # 3. Calculate interaction factors based on section class
        if pop.seclass in [1, 2]:  # Class 1 or 2 sections
            # k_yy - Formula from Table B.1
            term1 = 1 + (lamy - 0.2) * n_y
            term2 = 1 + 0.8 * n_y
            k_yy = C_my * min(term1, term2)

            # k_zz
            term1_z = 1 + (2 * lamz - 0.6) * n_z
            term2_z = 1 + 1.4 * n_z
            k_zz = C_mz * min(term1_z, term2_z)

            # k_yz for Class 1/2 sections
            k_yz = 0.6 * k_zz

            # k_zy for Class 1/2 sections
            if lamy < 0.4:
                k_zy = 0.6 * k_yy
            else:
                denominator = max(C_mLT - 0.25, 0.01)
                k_zy = 1 - (0.1 * lamz) / denominator * n_z
                k_zy = max(k_zy, 0.6 * k_yy)

        elif pop.seclass == 3:  # Class 3 sections
            # k_yy - same as Class 1/2 but with limits
            term1 = 1 + (0.6*lamy) * n_y
            term2 = 1 + 0.6 * n_y
            k_yy = C_my * min(term1, term2)

            # k_zz
            term1_z = 1 + (0.6 * lamz) * n_z
            term2_z = 1 + 0.6 * n_z
            k_zz = C_mz * min(term1_z, term2_z)

            # For Class 3 sections, k_yz and k_zy from Table B.1
            if lamz < 0.4:
                k_yz = k_zz
            else:
                k_yz = k_zz

            if lamy < 0.4:
                k_zy = 0.8 * k_zz
            else:
                if lamz < 0.4:
                    k_zy = 0.6 + lamz
                else:
                    denominator = max(C_mLT - 0.25, 0.01)
                    k_zy = 1 - (0.05 * lamz) / denominator * n_z
                    k_zy = max(k_zy, 0.6 * k_zz)

        else:  # Class 4 sections - use conservative approach
            k_yy = 1 + 0.6 * n_y
            k_zz = 1 + 0.6 * n_z
            k_yz = 0.6 * k_zz
            k_zy = 0.6 * k_yy

        # Ensure factors are within reasonable bounds
        k_yy = max(0.1, min(k_yy, 2.0))
        k_zz = max(0.1, min(k_zz, 2.0))
        k_yz = max(0.1, min(k_yz, 2.0))
        k_zy = max(0.1, min(k_zy, 2.0))

        # 4. Apply interaction checks (EC3 Eq. 6.61 and 6.62)
        # Eq. 6.61 — major-axis bending with LTB
        util_y = (Ned / Nbrd) + k_yy * (Myed / Mbrd) + k_yz * (Mzed / Mzrd)

        # Eq. 6.62 — minor-axis bending
        util_z = (Ned / Nbrd) + k_zy * (Myed / Mbrd) + k_zz * (Mzed / Mzrd)

        # Governing utilisation from EC3 Eq. 6.61 and 6.62 only
        U = max(util_y, util_z)
        # ===================== END MODIFIED SECTION =====================

        # -------------------------
        # CHECK
        # -------------------------
        if U <= 1.0:
            return {
                "Designation": pop.designation,
                "class": pop.seclass,
                "N_b_Rd": Nbrd,
                "M_b_Rd": Mbrd,
                "M_z_Rd": Mzrd,
                "utilisation": U,
                "chi_y": chiy,
                "chi_z": chiz,
                "chi_LT": chi_LT,
                # Add interaction factors to output for verification
                "k_yy": k_yy,
                "k_zz": k_zz,
                "k_yz": k_yz,
                "k_zy": k_zy,
                "C_my": C_my,
                "C_mz": C_mz,
                "util_y": util_y,
                "util_z": util_z
            }

        # Otherwise try next section
        i += 1

        # Safety break
        if i > ex.max_row:
            raise ValueError("No suitable section found.")




