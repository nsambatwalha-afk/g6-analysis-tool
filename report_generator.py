"""
report_generator.py
Generates formatted Excel results sheets for each design module.
Each public function returns a BytesIO buffer ready for st.download_button.
"""

import io
import math
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────
_TITLE_FILL   = PatternFill("solid", fgColor="1F4E79")
_TITLE_FONT   = Font(bold=True, size=14, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="BDD7EE")
_SECTION_FONT = Font(bold=True, size=11)
_LABEL_FONT   = Font(bold=True, size=10)
_NORMAL_FONT  = Font(size=10)
_PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")
_PASS_FONT    = Font(bold=True, size=10, color="276221")
_FAIL_FONT    = Font(bold=True, size=10, color="9C0006")
_THIN = Side(style="thin")
_BOX  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COL_WIDTHS = [4, 30, 50, 20]  # index, label, equation, value


# ─────────────────────────────────────────────
# Low-level helpers
# ─────────────────────────────────────────────

def _new_wb(title: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Design Report"

    # Column widths
    for ci, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Title row
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = title
    c.font = _TITLE_FONT
    c.fill = _TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    return wb, ws, 2  # workbook, sheet, next_row


def _section(ws, row: int, label: str) -> int:
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font = _SECTION_FONT
    c.fill = _SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    return row + 1


def _step(ws, row: int, step_no, label: str, equation: str, value: str = "") -> int:
    ws.cell(row=row, column=1, value=step_no).font  = _LABEL_FONT
    ws.cell(row=row, column=2, value=label).font    = _LABEL_FONT
    ws.cell(row=row, column=3, value=equation).font = _NORMAL_FONT
    ws.cell(row=row, column=4, value=value).font    = _NORMAL_FONT
    return row + 1


def _blank(ws, row: int) -> int:
    return row + 1


def _result_row(ws, row: int, label: str, value: str, passed: bool = True) -> int:
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font  = _PASS_FONT if passed else _FAIL_FONT
    c.fill  = _PASS_FILL if passed else _FAIL_FILL
    c.alignment = Alignment(horizontal="right")
    v = ws.cell(row=row, column=4, value=value)
    v.font  = _PASS_FONT if passed else _FAIL_FONT
    v.fill  = _PASS_FILL if passed else _FAIL_FILL
    return row + 1


def _wb_bytes(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fmt(val, dp=3):
    """Format a numeric value with dp decimal places."""
    try:
        return f"{float(val):.{dp}f}"
    except Exception:
        return str(val)


# ─────────────────────────────────────────────
# 1. Matrix Multiplication
# ─────────────────────────────────────────────

def matrix_mult_report(m1: np.matrix, m2: np.matrix, result: np.matrix) -> io.BytesIO:
    wb, ws, row = _new_wb("Matrix Multiplication — Results Sheet")

    row = _section(ws, row, "INPUT MATRICES")
    row = _step(ws, row, "", "Matrix A dimensions",
                f"{m1.shape[0]} rows × {m1.shape[1]} columns")
    row = _step(ws, row, "", "Matrix B dimensions",
                f"{m2.shape[0]} rows × {m2.shape[1]} columns")
    row = _blank(ws, row)

    # Matrix A
    row = _section(ws, row, "Matrix A")
    for r in range(m1.shape[0]):
        vals = "  ".join(_fmt(m1[r, c]) for c in range(m1.shape[1]))
        row = _step(ws, row, f"Row {r+1}", "", vals)

    row = _blank(ws, row)

    # Matrix B
    row = _section(ws, row, "Matrix B")
    for r in range(m2.shape[0]):
        vals = "  ".join(_fmt(m2[r, c]) for c in range(m2.shape[1]))
        row = _step(ws, row, f"Row {r+1}", "", vals)

    row = _blank(ws, row)

    # Computation
    row = _section(ws, row, "COMPUTATION — Element-by-Element")
    row = _step(ws, row, "", "General formula", "C[i,j] = Σ  A[i,k] × B[k,j]   for k = 1 … N")
    row = _blank(ws, row)

    rows_a, cols_a = m1.shape
    _, cols_b = m2.shape
    for i in range(rows_a):
        for j in range(cols_b):
            terms = [f"A[{i+1},{k+1}]×B[{k+1},{j+1}] = {_fmt(m1[i,k])}×{_fmt(m2[k,j])}"
                     for k in range(cols_a)]
            eq = " + ".join(terms)
            row = _step(ws, row, f"C[{i+1},{j+1}]",
                        eq,
                        _fmt(result[i, j]))

    row = _blank(ws, row)

    # Result matrix
    row = _section(ws, row, "RESULT MATRIX  C = A × B")
    for r in range(result.shape[0]):
        vals = "  ".join(_fmt(result[r, c]) for c in range(result.shape[1]))
        row = _step(ws, row, f"Row {r+1}", "", vals)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# 2. Gauss-Jordan Elimination
# ─────────────────────────────────────────────

def gauss_jordan_report(aug_input: np.matrix) -> io.BytesIO:
    """
    Re-runs the Gauss-Jordan elimination, capturing every step.
    aug_input should be the original augmented matrix [A|b].
    """
    wb, ws, row = _new_wb("Gauss-Jordan Elimination — Results Sheet")

    arg = aug_input.copy().astype(float)
    matrix_shape = arg.shape
    rows = matrix_shape[0]
    cols = matrix_shape[1] - 1

    def mat_str(m):
        lines = []
        for r in range(m.shape[0]):
            vals = "  ".join(_fmt(m[r, c]) for c in range(m.shape[1]))
            lines.append(vals)
        return " | ".join(lines)

    # Input
    row = _section(ws, row, "INPUT AUGMENTED MATRIX [A | b]")
    for r in range(arg.shape[0]):
        vals = "  ".join(_fmt(arg[r, c]) for c in range(arg.shape[1]))
        row = _step(ws, row, f"Row {r+1}", "", vals)

    row = _blank(ws, row)

    # ── Forward elimination ──
    row = _section(ws, row, "FORWARD ELIMINATION")
    step_no = 1
    for j in range(cols):
        i = rows - 1
        while i > j:
            if arg[j, j] == 0:
                row = _step(ws, row, f"Step {step_no}",
                            "Diagonal element is zero — elimination not suitable for this method.",
                            "ABORT")
                return _wb_bytes(wb)
            factor = arg[i, j] / arg[j, j]
            eq = (f"R{i+1} ← R{i+1} - ({_fmt(arg[i,j])} / {_fmt(arg[j,j])}) × R{j+1}  "
                  f"[factor = {_fmt(factor)}]")
            arg[i] = -factor * arg[j] + arg[i]
            row = _step(ws, row, f"Step {step_no}", eq, "")
            # Show updated matrix
            for r in range(arg.shape[0]):
                vals = "  ".join(_fmt(arg[r, c]) for c in range(arg.shape[1]))
                row = _step(ws, row, "", f"  Row {r+1}", vals)
            row = _blank(ws, row)
            step_no += 1
            i -= 1

    # ── Back substitution ──
    row = _section(ws, row, "BACK SUBSTITUTION")
    j = cols - 1
    while j > 0:
        i = 0
        while i < j:
            if arg[j, j] == 0:
                row = _step(ws, row, f"Step {step_no}",
                            "Diagonal element is zero — cannot continue.", "ABORT")
                return _wb_bytes(wb)
            factor = arg[i, j] / arg[j, j]
            eq = (f"R{i+1} ← R{i+1} - ({_fmt(arg[i,j])} / {_fmt(arg[j,j])}) × R{j+1}  "
                  f"[factor = {_fmt(factor)}]")
            arg[i] = -factor * arg[j] + arg[i]
            row = _step(ws, row, f"Step {step_no}", eq, "")
            for r in range(arg.shape[0]):
                vals = "  ".join(_fmt(arg[r, c]) for c in range(arg.shape[1]))
                row = _step(ws, row, "", f"  Row {r+1}", vals)
            row = _blank(ws, row)
            step_no += 1
            i += 1
        j -= 1

    # ── Normalize ──
    row = _section(ws, row, "NORMALISATION (Divide each row by its pivot)")
    for i in range(rows):
        pivot = arg[i, i]
        eq = f"R{i+1} ← R{i+1} / {_fmt(pivot)}"
        arg[i] = arg[i] / pivot
        row = _step(ws, row, f"Step {step_no}", eq, "")
        step_no += 1

    row = _blank(ws, row)

    # ── Solution ──
    row = _section(ws, row, "SOLUTION")
    for i in range(rows):
        row = _step(ws, row, f"x{i+1}", f"x{i+1} =", _fmt(arg[i, cols]))

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# Shared helper: look up grade properties
# ─────────────────────────────────────────────

def get_grade_props(grade: str):
    """Returns (fy, fu) for the given grade string."""
    try:
        wb = openpyxl.load_workbook("grades.xlsx").active
    except FileNotFoundError:
        raise FileNotFoundError(
            "grades.xlsx not found. Ensure the grade properties file is present "
            "in the working directory."
        )
    for i in range(2, 10):
        if wb.cell(row=i, column=1).value == grade:
            return float(wb.cell(row=i, column=2).value), float(wb.cell(row=i, column=3).value)
    raise ValueError(f"Grade '{grade}' not found in grades.xlsx")


# ─────────────────────────────────────────────
# 3. Tension Member Design
# ─────────────────────────────────────────────

def tension_design_report(
    force_N: float,          # design force in N
    grade: str,
    shapeten: str,
    section,                 # tuple returned by table_reader
    jointing: str,
    nh: float = 0,
    d: float = 0,
    stag: bool = False,
    s: float = 0,
    p: float = 0,
    ngs: float = 0
) -> io.BytesIO:
    wb, ws, row = _new_wb("Tension Member Design — Results Sheet (EC3)")

    fy, fu = get_grade_props(grade)
    agross = section[0]
    size   = section[1]
    t      = section[2]

    # Net area
    if jointing == "weld":
        anet = agross
        net_eq = "A_net = A_gross  (welded connection)"
        net_sub = f"A_net = {_fmt(agross)} mm²"
    elif jointing == "bolt" and not stag:
        anet = agross - nh * t * (d + 2.0)
        net_eq = "A_net = A_gross − n_h × t × (d + 2)"
        net_sub = f"A_net = {_fmt(agross)} − {_fmt(nh)} × {_fmt(t)} × ({_fmt(d)} + 2) = {_fmt(anet)} mm²"
    else:  # staggered
        anet = agross - nh * t * (d + 2.0) + ngs * (((s ** 2) * t) / (4 * p))
        net_eq = "A_net = A_gross − n_h×t×(d+2) + n_gs×(s²×t)/(4p)"
        net_sub = (f"A_net = {_fmt(agross)} − {_fmt(nh)}×{_fmt(t)}×({_fmt(d)}+2)"
                   f" + {_fmt(ngs)}×({_fmt(s)}²×{_fmt(t)})/(4×{_fmt(p)}) = {_fmt(anet)} mm²")

    npl = agross * fy / 1.0          # N
    nu  = 0.9 * anet * fu / 1.25     # N
    ntrd = min(npl, nu)
    passed = force_N <= ntrd

    # ── INPUTS ──
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Design Axial Force",     f"N_Ed = {_fmt(force_N/1000, 2)} kN  ({_fmt(force_N)} N)")
    row = _step(ws, row, "·", "Steel Grade",             grade)
    row = _step(ws, row, "·", "Member Shape",            shapeten)
    row = _step(ws, row, "·", "Connection Type",         jointing)
    if jointing == "bolt":
        row = _step(ws, row, "·", "Number of bolt holes",   str(int(nh)))
        row = _step(ws, row, "·", "Bolt diameter d",         f"{_fmt(d)} mm")
        if stag:
            row = _step(ws, row, "·", "Stagger spacing s",   f"{_fmt(s)} mm")
            row = _step(ws, row, "·", "Pitch p",             f"{_fmt(p)} mm")
            row = _step(ws, row, "·", "No. stagger lines",   str(int(ngs)))
    row = _blank(ws, row)

    # ── MATERIAL PROPERTIES ──
    row = _section(ws, row, "MATERIAL PROPERTIES (from grades.xlsx)")
    row = _step(ws, row, "1", "Yield strength",
                f"f_y ({grade})", f"{_fmt(fy)} MPa")
    row = _step(ws, row, "2", "Ultimate strength",
                f"f_u ({grade})", f"{_fmt(fu)} MPa")
    row = _blank(ws, row)

    # ── STEP-BY-STEP DESIGN ──
    row = _section(ws, row, "STEP-BY-STEP DESIGN CALCULATIONS")

    row = _step(ws, row, "Step 1",
                "Required gross area",
                "A_req = N_Ed / f_y",
                f"{_fmt(force_N / fy)} mm²")

    row = _step(ws, row, "Step 2",
                f"Section selected: {size}",
                "A_gross (from section table)",
                f"{_fmt(agross)} mm²")
    row = _step(ws, row, "",
                "Plate / leg thickness t",
                "",
                f"{_fmt(t)} mm")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 3",
                "Net area calculation",
                net_eq, "")
    row = _step(ws, row, "",
                "Substitution",
                net_sub,
                f"{_fmt(anet)} mm²")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 4",
                "Plastic resistance (yielding of gross section)",
                "N_pl,Rd = A_gross × f_y / γ_M0     [γ_M0 = 1.0]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"N_pl,Rd = {_fmt(agross)} × {_fmt(fy)} / 1.0",
                f"{_fmt(npl/1000, 2)} kN")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 5",
                "Ultimate resistance (fracture of net section)",
                "N_u,Rd = 0.9 × A_net × f_u / γ_M2     [γ_M2 = 1.25]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"N_u,Rd = 0.9 × {_fmt(anet)} × {_fmt(fu)} / 1.25",
                f"{_fmt(nu/1000, 2)} kN")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 6",
                "Design tensile resistance (governing)",
                "N_t,Rd = min(N_pl,Rd , N_u,Rd)",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"N_t,Rd = min({_fmt(npl/1000, 2)}, {_fmt(nu/1000, 2)}) kN",
                f"{_fmt(ntrd/1000, 2)} kN")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 7",
                "Capacity check",
                "N_Ed ≤ N_t,Rd",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"{_fmt(force_N/1000, 2)} kN  ≤  {_fmt(ntrd/1000, 2)} kN",
                "PASS ✓" if passed else "FAIL ✗")

    row = _blank(ws, row)
    row = _result_row(ws, row,
                      f"Section: {size}  |  N_t,Rd = {_fmt(ntrd/1000, 2)} kN  |  N_Ed = {_fmt(force_N/1000, 2)} kN",
                      "PASS ✓" if passed else "FAIL ✗",
                      passed)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# 4. Compression Member Design
# ─────────────────────────────────────────────

def compression_design_report(
    force_N: float,
    L: float,
    grade: str,
    shapecomp: str,
    section: dict
) -> io.BytesIO:
    wb, ws, row = _new_wb("Compression Member Design — Results Sheet (EC3)")

    fy, _ = get_grade_props(grade)
    E = 210000  # MPa

    if shapecomp in ("UC", "UB"):
        alpha = 0.34
    elif shapecomp == "CHS":
        alpha = 0.21
    elif shapecomp == "Angle":
        alpha = 0.49
    else:
        alpha = 0.49

    A       = section["area"]
    size    = section["size"]
    I       = section["I"]
    r       = section["r"]
    axis    = section["axis"]
    lam_bar = section["lambda_bar"]
    chi     = section["chi"]
    NbRd    = section["NbRd"]

    Ncr  = (math.pi ** 2 * E * I) / (L ** 2)
    phi  = 0.5 * (1 + alpha * (lam_bar - 0.2) + lam_bar ** 2)

    # ── INPUTS ──
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Design Axial Force (compression)",
                f"N_Ed = {_fmt(force_N/1000, 2)} kN  ({_fmt(force_N)} N)")
    row = _step(ws, row, "·", "Member Length",    f"L = {_fmt(L)} mm")
    row = _step(ws, row, "·", "Steel Grade",      grade)
    row = _step(ws, row, "·", "Section Shape",    shapecomp)
    row = _blank(ws, row)

    # ── MATERIAL / GEOMETRY ──
    row = _section(ws, row, "MATERIAL & SECTION PROPERTIES")
    row = _step(ws, row, "1", "Yield strength",       f"f_y ({grade})", f"{_fmt(fy)} MPa")
    row = _step(ws, row, "2", "Elastic modulus",      "E",              f"210 000 MPa")
    row = _step(ws, row, "3", "Imperfection factor α (buckling curve)",
                f"shape = {shapecomp}", f"α = {_fmt(alpha)}")
    row = _step(ws, row, "4", f"Section selected",    size, "")
    row = _step(ws, row, "", "Gross area A",          "", f"{_fmt(A)} mm²")
    row = _step(ws, row, "", f"Second moment of area (about {axis}-axis)  I",
                "", f"{_fmt(I)} mm⁴")
    row = _step(ws, row, "", "Radius of gyration r",  "", f"{_fmt(r)} mm")
    row = _blank(ws, row)

    # ── CALCULATIONS ──
    row = _section(ws, row, "STEP-BY-STEP CALCULATIONS")

    row = _step(ws, row, "Step 1",
                "Elastic critical buckling force",
                "N_cr = π² × E × I / L²",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"N_cr = π² × {_fmt(E)} × {_fmt(I)} / {_fmt(L)}²",
                f"{_fmt(Ncr/1000, 2)} kN")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 2",
                "Non-dimensional slenderness",
                "λ̄ = √(A × f_y / N_cr)",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"λ̄ = √({_fmt(A)} × {_fmt(fy)} / {_fmt(Ncr, 1)})",
                f"λ̄ = {_fmt(lam_bar)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 3",
                "Buckling curve factor φ",
                "φ = 0.5 × [1 + α(λ̄ − 0.2) + λ̄²]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"φ = 0.5 × [1 + {_fmt(alpha)}×({_fmt(lam_bar)}−0.2) + {_fmt(lam_bar)}²]",
                f"φ = {_fmt(phi)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 4",
                "Reduction factor χ",
                "χ = 1 / [φ + √(φ² − λ̄²)]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"χ = 1 / [{_fmt(phi)} + √({_fmt(phi)}² − {_fmt(lam_bar)}²)]",
                f"χ = {_fmt(chi)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 5",
                "Buckling resistance",
                "N_b,Rd = χ × A × f_y / γ_M1     [γ_M1 = 1.0]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"N_b,Rd = {_fmt(chi)} × {_fmt(A)} × {_fmt(fy)} / 1.0",
                f"{_fmt(NbRd/1000, 2)} kN")
    row = _blank(ws, row)

    passed = force_N <= NbRd
    row = _step(ws, row, "Step 6",
                "Capacity check",
                "N_Ed ≤ N_b,Rd",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"{_fmt(force_N/1000, 2)} kN ≤ {_fmt(NbRd/1000, 2)} kN",
                "PASS ✓" if passed else "FAIL ✗")

    row = _blank(ws, row)
    util = force_N / NbRd * 100
    row = _result_row(ws, row,
                      f"Section: {size}  |  N_b,Rd = {_fmt(NbRd/1000, 2)} kN  "
                      f"|  χ = {_fmt(chi)}  |  Utilisation = {_fmt(util, 1)} %",
                      "PASS ✓" if passed else "FAIL ✗",
                      passed)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# 5. Restrained Beam Design
# ─────────────────────────────────────────────

def restrained_beam_report(M: float, V: float, grade: str, result: dict) -> io.BytesIO:
    """
    M in kNm, V in kN.
    result is the dict returned by truss_analysis.restrained_beam().
    Re-reads section properties from UB-2.xlsx to show step-by-step.
    """
    wb, ws, row = _new_wb("Restrained Beam Design — Results Sheet (EC3)")

    fy, _ = get_grade_props(grade)
    gamma_M0 = 1.0

    # Re-read section props for the chosen size
    try:
        ex = openpyxl.load_workbook("UB-2.xlsx").active
    except FileNotFoundError:
        raise FileNotFoundError(
            "UB-2.xlsx not found. Ensure the UB section properties file is present "
            "in the working directory."
        )
    size = result["Size"]
    Wpl = Av = Aw = tw = h = tf = None
    for i in range(2, ex.max_row + 1):
        if ex.cell(row=i, column=1).value == size:
            Wyy = float(ex.cell(row=i, column=8).value) * 1000
            Wzz = float(ex.cell(row=i, column=9).value) * 1000
            Wpl = max(Wyy, Wzz)
            tw  = float(ex.cell(row=i, column=15).value)
            h   = float(ex.cell(row=i, column=16).value)
            tf  = float(ex.cell(row=i, column=17).value)
            hw  = h - 2 * tf
            Av  = hw * tw
            Aw  = hw * tw
            break

    Vpl_Rd = result["V_Rd (kN)"]   # kN (already computed)
    M_Rd   = result["M_Rd (kNm)"]  # kNm
    util   = result["Utilization (%)"]

    high_shear = V > 0.5 * Vpl_Rd

    # ── INPUTS ──
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Design Bending Moment",   f"M_Ed = {_fmt(M, 2)} kNm")
    row = _step(ws, row, "·", "Design Shear Force",      f"V_Ed = {_fmt(V, 2)} kN")
    row = _step(ws, row, "·", "Steel Grade",             grade)
    row = _step(ws, row, "·", "Beam Type",               "Restrained (full lateral restraint)")
    row = _blank(ws, row)

    # ── MATERIAL ──
    row = _section(ws, row, "MATERIAL PROPERTIES")
    row = _step(ws, row, "1", "Yield strength", f"f_y ({grade})", f"{_fmt(fy)} MPa")
    row = _step(ws, row, "2", "γ_M0", "", "1.0")
    row = _blank(ws, row)

    # ── SECTION ──
    row = _section(ws, row, f"SELECTED SECTION: {size}")
    if Wpl is not None:
        row = _step(ws, row, "·", "Plastic section modulus  W_pl", "", f"{_fmt(Wpl, 1)} mm³")
        row = _step(ws, row, "·", "Web thickness  t_w",            "", f"{_fmt(tw)} mm")
        row = _step(ws, row, "·", "Overall depth  h",              "", f"{_fmt(h)} mm")
        row = _step(ws, row, "·", "Flange thickness  t_f",         "", f"{_fmt(tf)} mm")
        row = _step(ws, row, "·", "Shear area  A_v = h_w × t_w",  "", f"{_fmt(Av, 1)} mm²")
    row = _blank(ws, row)

    # ── CALCULATIONS ──
    row = _section(ws, row, "STEP-BY-STEP CALCULATIONS")

    row = _step(ws, row, "Step 1",
                "Shear resistance",
                "V_pl,Rd = A_v × f_y / (√3 × γ_M0)  [kN]",
                "")
    if Av is not None:
        row = _step(ws, row, "",
                    "Substitution",
                    f"V_pl,Rd = {_fmt(Av, 1)} × {_fmt(fy)} / (√3 × 1.0) / 1000",
                    f"{_fmt(Vpl_Rd, 2)} kN")
    row = _step(ws, row, "",
                "Check: V_Ed ≤ V_pl,Rd",
                f"{_fmt(V, 2)} kN ≤ {_fmt(Vpl_Rd, 2)} kN",
                "PASS ✓" if V <= Vpl_Rd else "FAIL ✗")
    row = _blank(ws, row)

    if not high_shear:
        row = _step(ws, row, "Step 2",
                    "Low shear (V_Ed / V_pl,Rd ≤ 0.5) — no reduction",
                    f"V_Ed / V_pl,Rd = {_fmt(V/Vpl_Rd, 3)} ≤ 0.5",
                    "No reduction")
        row = _step(ws, row, "",
                    "Moment resistance",
                    "M_pl,Rd = W_pl × f_y / γ_M0  [kNm]",
                    "")
        if Wpl is not None:
            row = _step(ws, row, "",
                        "Substitution",
                        f"M_pl,Rd = {_fmt(Wpl, 1)} × {_fmt(fy)} / 1.0 / 1e6",
                        f"{_fmt(M_Rd, 2)} kNm")
    else:
        row = _step(ws, row, "Step 2",
                    "High shear — compute reduction factor ρ",
                    "ρ = (2 × V_Ed / V_pl,Rd − 1)²",
                    "")
        rho = (2 * V / Vpl_Rd - 1) ** 2
        row = _step(ws, row, "",
                    "Substitution",
                    f"ρ = (2 × {_fmt(V, 2)} / {_fmt(Vpl_Rd, 2)} − 1)²",
                    f"ρ = {_fmt(rho)}")
        row = _blank(ws, row)
        row = _step(ws, row, "Step 3",
                    "Reduced moment resistance",
                    "M_V,Rd = [W_pl − (ρ × A_w² / (4 × t_w))] × f_y / γ_M0",
                    "")
        if Wpl is not None and Aw is not None:
            row = _step(ws, row, "",
                        "Substitution",
                        f"M_V,Rd = [{_fmt(Wpl,1)} − ({_fmt(rho)} × {_fmt(Aw,1)}² / (4×{_fmt(tw)}))] × {_fmt(fy)} / 1.0 / 1e6",
                        f"{_fmt(M_Rd, 2)} kNm")

    row = _blank(ws, row)
    row = _step(ws, row, "Step 3" if not high_shear else "Step 4",
                "Moment capacity check",
                "M_Ed ≤ M_Rd",
                "")
    passed = M <= M_Rd
    row = _step(ws, row, "",
                "Substitution",
                f"{_fmt(M, 2)} kNm ≤ {_fmt(M_Rd, 2)} kNm",
                "PASS ✓" if passed else "FAIL ✗")

    row = _blank(ws, row)
    row = _result_row(ws, row,
                      f"Section: {size}  |  M_Rd = {_fmt(M_Rd, 2)} kNm  "
                      f"|  V_Rd = {_fmt(Vpl_Rd, 2)} kN  |  Utilisation = {_fmt(util, 1)} %",
                      "PASS ✓" if passed else "FAIL ✗",
                      passed)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# 6. Unrestrained Beam Design
# ─────────────────────────────────────────────

def unrestrained_beam_report(
    M: float,
    V: float,
    L: float,
    grade: str,
    condition: str,
    endcondition: str,
    result: dict
) -> io.BytesIO:
    """
    M in kNm, V in kN, L in mm.
    result is the dict returned by truss_analysis.unrestrained_beam().
    """
    wb, ws, row = _new_wb("Unrestrained Beam Design — Results Sheet (EC3 LTB)")

    fy, _ = get_grade_props(grade)
    E = 210000
    gamma_M1 = 1.0

    size = result["Size"]

    # Re-read section
    try:
        ex = openpyxl.load_workbook("UB-2.xlsx").active
    except FileNotFoundError:
        raise FileNotFoundError(
            "UB-2.xlsx not found. Ensure the UB section properties file is present "
            "in the working directory."
        )
    sec = {}
    for i in range(2, ex.max_row + 1):
        if ex.cell(row=i, column=1).value == size:
            sec["Wyy"] = float(ex.cell(row=i, column=8).value) * 1000
            sec["Iy"]  = float(ex.cell(row=i, column=2).value) * 10000
            sec["Iz"]  = float(ex.cell(row=i, column=3).value) * 10000
            sec["tw"]  = float(ex.cell(row=i, column=15).value)
            sec["h"]   = float(ex.cell(row=i, column=16).value)
            sec["tf"]  = float(ex.cell(row=i, column=17).value)
            sec["b"]   = float(ex.cell(row=i, column=19).value)
            sec["iz"]  = float(ex.cell(row=i, column=5).value) * 10.0
            break

    Wpl = sec.get("Wyy", 0)
    h   = sec.get("h",  0)
    b   = sec.get("b",  0)
    tw  = sec.get("tw", 0)
    tf  = sec.get("tf", 0)
    iz  = sec.get("iz", 0)
    hw  = h - 2 * tf
    Av  = hw * tw
    Aw  = hw * tw

    k_map = {"Free": 1.0, "Partial": 0.85, "Full": 0.7, "Cantilever": 2.0}
    k = k_map.get(endcondition, 1.0)

    lamz    = (k * L) / iz if iz else 0
    laml    = math.pi * math.sqrt(E / fy)
    lamzba  = lamz / laml if laml else 0
    c1, u, vee, bew = 1.0, 0.9, 1.0, 1.0
    lamltb  = (1 / math.sqrt(c1)) * u * vee * lamzba * math.sqrt(bew)
    hoverb  = h / b if b else 0

    if condition == "Rolled":
        alt = 0.34 if hoverb <= 2 else 0.49
        phi = 0.5 * (1 + alt * (lamltb - 0.4) + 0.75 * (lamltb ** 2))
        chi = min(1 / (phi + math.sqrt(phi ** 2 - 0.75 * (lamltb ** 2))), 1.0)
    else:
        alt = 0.49 if hoverb <= 2 else 0.76
        phi = 0.5 * (1 + alt * (lamltb - 0.2) + (lamltb ** 2))
        chi = min(1 / (phi + math.sqrt(phi ** 2 - (lamltb ** 2))), 1.0)

    Vpl_Rd = result.get("Vpl_Rd", (Av * fy) / (math.sqrt(3) * gamma_M1) / 1000)

    high_shear = V > 0.5 * Vpl_Rd
    if high_shear:
        rho = ((2 * V / Vpl_Rd) - 1) ** 2
        W_red = Wpl - (rho * (Aw ** 2)) / (4 * tw)
    else:
        rho = 0.0
        W_red = Wpl

    Mbrd = result["Mb_Rd (kNm)"]
    util = result["Utilization (%)"]

    # ── INPUTS ──
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Design Bending Moment", f"M_Ed = {_fmt(M, 2)} kNm")
    row = _step(ws, row, "·", "Design Shear Force",    f"V_Ed = {_fmt(V, 2)} kN")
    row = _step(ws, row, "·", "Beam Length",            f"L = {_fmt(L)} mm")
    row = _step(ws, row, "·", "Steel Grade",            grade)
    row = _step(ws, row, "·", "Beam Condition",         condition)
    row = _step(ws, row, "·", "End Restraint",          endcondition)
    row = _blank(ws, row)

    # ── MATERIAL ──
    row = _section(ws, row, "MATERIAL PROPERTIES")
    row = _step(ws, row, "1", "Yield strength",  f"f_y ({grade})", f"{_fmt(fy)} MPa")
    row = _step(ws, row, "2", "Elastic modulus", "E",              "210 000 MPa")
    row = _step(ws, row, "3", "γ_M1",            "",               "1.0")
    row = _blank(ws, row)

    # ── SECTION ──
    row = _section(ws, row, f"SELECTED SECTION: {size}")
    row = _step(ws, row, "·", "Plastic section modulus W_pl", "", f"{_fmt(Wpl, 1)} mm³")
    row = _step(ws, row, "·", "Minor-axis radius of gyration i_z", "", f"{_fmt(iz)} mm")
    row = _step(ws, row, "·", "h / b ratio", f"{_fmt(h, 1)} / {_fmt(b, 1)}", f"{_fmt(hoverb, 3)}")
    row = _step(ws, row, "·", "Shear area A_v = h_w × t_w", "", f"{_fmt(Av, 1)} mm²")
    row = _blank(ws, row)

    # ── CALCULATIONS ──
    row = _section(ws, row, "STEP-BY-STEP CALCULATIONS")

    # Effective length factor
    row = _step(ws, row, "Step 1",
                "Effective length factor k  (from end restraint)",
                f"Restraint = '{endcondition}'",
                f"k = {_fmt(k)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 2",
                "Slenderness ratio (minor axis)",
                "λ_z = k × L / i_z",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"λ_z = {_fmt(k)} × {_fmt(L)} / {_fmt(iz)}",
                f"λ_z = {_fmt(lamz, 2)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 3",
                "Limiting slenderness  λ₁",
                "λ₁ = π × √(E / f_y)",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"λ₁ = π × √({_fmt(E)} / {_fmt(fy)})",
                f"λ₁ = {_fmt(laml, 2)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 4",
                "Normalised minor-axis slenderness  λ̄_z",
                "λ̄_z = λ_z / λ₁",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"λ̄_z = {_fmt(lamz, 2)} / {_fmt(laml, 2)}",
                f"λ̄_z = {_fmt(lamzba, 4)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 5",
                "LTB slenderness  λ̄_LT  (simplified method)",
                "λ̄_LT = (1/√C₁) × u × v × λ̄_z × √β_w   [C₁=1, u=0.9, v=1, β_w=1]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"λ̄_LT = (1/√{_fmt(c1)}) × {_fmt(u)} × {_fmt(vee)} × {_fmt(lamzba,4)} × √{_fmt(bew)}",
                f"λ̄_LT = {_fmt(lamltb, 4)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 6",
                f"Imperfection factor α_LT  (condition={condition}, h/b={_fmt(hoverb,2)})",
                f"α_LT = {_fmt(alt)}",
                "")
    row = _blank(ws, row)

    if condition == "Rolled":
        phi_eq  = "φ_LT = 0.5 × [1 + α_LT(λ̄_LT − 0.4) + 0.75 × λ̄_LT²]"
        chi_eq  = "χ_LT = 1 / [φ_LT + √(φ_LT² − 0.75 × λ̄_LT²)]  ≤ 1.0"
    else:
        phi_eq  = "φ_LT = 0.5 × [1 + α_LT(λ̄_LT − 0.2) + λ̄_LT²]"
        chi_eq  = "χ_LT = 1 / [φ_LT + √(φ_LT² − λ̄_LT²)]  ≤ 1.0"

    row = _step(ws, row, "Step 7", "φ_LT", phi_eq, "")
    row = _step(ws, row, "",
                "Substitution",
                f"φ_LT = {_fmt(phi, 4)}",
                "")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 8", "LTB reduction factor χ_LT", chi_eq, "")
    row = _step(ws, row, "",
                "Substitution",
                f"χ_LT = {_fmt(chi, 4)}",
                "")
    row = _blank(ws, row)

    row = _step(ws, row, "Step 9",
                "Shear resistance",
                "V_pl,Rd = A_v × f_y / (√3 × γ_M1)  [kN]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"V_pl,Rd = {_fmt(Av,1)} × {_fmt(fy)} / (√3 × 1.0) / 1000",
                f"{_fmt(Vpl_Rd, 2)} kN")
    row = _step(ws, row, "",
                "Shear check: V_Ed ≤ V_pl,Rd",
                f"{_fmt(V, 2)} kN ≤ {_fmt(Vpl_Rd, 2)} kN",
                "PASS ✓" if V <= Vpl_Rd else "FAIL ✗")
    row = _blank(ws, row)

    if high_shear:
        row = _step(ws, row, "Step 10",
                    "High shear — reduction factor ρ",
                    "ρ = (2 × V_Ed / V_pl,Rd − 1)²",
                    "")
        row = _step(ws, row, "",
                    "Substitution",
                    f"ρ = (2 × {_fmt(V, 2)} / {_fmt(Vpl_Rd, 2)} − 1)²",
                    f"ρ = {_fmt(rho)}")
        row = _step(ws, row, "Step 11",
                    "Reduced plastic modulus",
                    "W_red = W_pl − (ρ × A_w² / (4 × t_w))",
                    "")
        row = _step(ws, row, "",
                    "Substitution",
                    f"W_red = {_fmt(Wpl,1)} − ({_fmt(rho)} × {_fmt(Aw,1)}² / (4 × {_fmt(tw)}))",
                    f"W_red = {_fmt(W_red, 1)} mm³")
        step_lbl = "Step 12"
    else:
        row = _step(ws, row, "Step 10",
                    "Low shear — no reduction to moment capacity",
                    f"V_Ed / V_pl,Rd = {_fmt(V/Vpl_Rd, 3)} ≤ 0.5  →  W_red = W_pl",
                    "")
        step_lbl = "Step 11"

    row = _blank(ws, row)
    row = _step(ws, row, step_lbl,
                "Design buckling resistance moment",
                "M_b,Rd = χ_LT × W_red × f_y / γ_M1  [kNm]",
                "")
    row = _step(ws, row, "",
                "Substitution",
                f"M_b,Rd = {_fmt(chi,4)} × {_fmt(W_red,1)} × {_fmt(fy)} / 1.0 / 1e6",
                f"{_fmt(Mbrd, 2)} kNm")
    row = _blank(ws, row)

    passed = M <= Mbrd
    row = _step(ws, row, "Final check",
                "M_Ed ≤ M_b,Rd",
                f"{_fmt(M, 2)} kNm ≤ {_fmt(Mbrd, 2)} kNm",
                "PASS ✓" if passed else "FAIL ✗")

    row = _blank(ws, row)
    row = _result_row(ws, row,
                      f"Section: {size}  |  M_b,Rd = {_fmt(Mbrd, 2)} kNm  "
                      f"|  χ_LT = {_fmt(chi, 3)}  |  Utilisation = {_fmt(util, 1)} %",
                      "PASS ✓" if passed else "FAIL ✗",
                      passed)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# 7. Beam-Column Design
# ─────────────────────────────────────────────

def beam_column_report(
    L: float,
    Ned: float,
    Mzed: float,
    Myed: float,
    shape: str,
    C1: float,
    grade: str,
    endcondition,
    all_axis_similar: bool,
    result: dict
) -> io.BytesIO:
    wb, ws, row = _new_wb("Beam-Column Design — Results Sheet (EC3 Cl.6.3.3)")

    fy, _ = get_grade_props(grade)
    E = 210000
    G = 81000

    designation = result["Designation"]
    sec_class   = result["class"]
    Nbrd        = result["N_b_Rd"]
    Mbrd        = result["M_b_Rd"]
    Mzrd        = result["M_z_Rd"]
    U           = result["utilisation"]
    chiy        = result["chi_y"]
    chiz        = result["chi_z"]
    chiLT       = result["chi_LT"]
    k_yy        = result["k_yy"]
    k_zz        = result["k_zz"]
    k_yz        = result["k_yz"]
    k_zy        = result["k_zy"]
    C_my        = result["C_my"]
    C_mz        = result["C_mz"]
    util_y      = result["util_y"]
    util_z      = result["util_z"]

    # Effective lengths
    if all_axis_similar:
        Lcry = _eff_L(L, endcondition)
        Lcrz = Lcry
        ec_str = endcondition
    else:
        Lcrz = _eff_L(L, endcondition[0])
        Lcry = _eff_L(L, endcondition[1])
        ec_str = f"Z-axis: {endcondition[0]}, Y-axis: {endcondition[1]}"

    # ── INPUTS ──
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Member Length",         f"L = {_fmt(L)} mm")
    row = _step(ws, row, "·", "Design Axial Force",    f"N_Ed = {_fmt(Ned/1000, 2)} kN")
    row = _step(ws, row, "·", "Design Moment z-axis",  f"M_z,Ed = {_fmt(Mzed/1e6, 2)} kNm")
    row = _step(ws, row, "·", "Design Moment y-axis",  f"M_y,Ed = {_fmt(Myed/1e6, 2)} kNm")
    row = _step(ws, row, "·", "Section shape",         shape)
    row = _step(ws, row, "·", "Steel Grade",           grade)
    row = _step(ws, row, "·", "End conditions",        ec_str)
    row = _step(ws, row, "·", "C₁ (moment correction factor)", f"{_fmt(C1)}")
    row = _blank(ws, row)

    # ── SECTION / CLASS ──
    row = _section(ws, row, f"SELECTED SECTION: {designation}  —  Class {sec_class}")
    row = _step(ws, row, "·", "Cross-section class",   str(sec_class))
    row = _blank(ws, row)

    # ── EFFECTIVE LENGTHS ──
    row = _section(ws, row, "EFFECTIVE LENGTHS")
    row = _step(ws, row, "1", "Effective length about y-axis  L_cr,y",
                f"End condition: {endcondition if all_axis_similar else endcondition[1]}",
                f"L_cr,y = {_fmt(Lcry)} mm")
    row = _step(ws, row, "2", "Effective length about z-axis  L_cr,z",
                f"End condition: {endcondition if all_axis_similar else endcondition[0]}",
                f"L_cr,z = {_fmt(Lcrz)} mm")
    row = _blank(ws, row)

    # ── COMPRESSION CHECKS ──
    row = _section(ws, row, "FLEXURAL BUCKLING — χ_y and χ_z")
    row = _step(ws, row, "1", "Buckling resistance N_b,Rd",
                "N_b,Rd = χ × A × f_y    (χ = min(χ_y, χ_z))",
                f"{_fmt(Nbrd/1000, 2)} kN")
    row = _step(ws, row, "", "χ_y",  "", f"{_fmt(chiy, 4)}")
    row = _step(ws, row, "", "χ_z",  "", f"{_fmt(chiz, 4)}")
    row = _blank(ws, row)

    # ── LTB ──
    row = _section(ws, row, "LATERAL TORSIONAL BUCKLING — χ_LT")
    row = _step(ws, row, "1", "M_b,Rd = χ_LT × W_y × f_y",
                f"χ_LT = {_fmt(chiLT, 4)}",
                f"{_fmt(Mbrd/1e6, 2)} kNm")
    row = _blank(ws, row)

    # ── MINOR AXIS ──
    row = _section(ws, row, "MINOR AXIS BENDING RESISTANCE")
    row = _step(ws, row, "1", "M_z,Rd = W_z × f_y",
                "",
                f"{_fmt(Mzrd/1e6, 2)} kNm")
    row = _blank(ws, row)

    # ── INTERACTION FACTORS ──
    row = _section(ws, row, "INTERACTION FACTORS (EC3 Annex B)")
    row = _step(ws, row, "·", "C_my", "", f"{_fmt(C_my, 4)}")
    row = _step(ws, row, "·", "C_mz", "", f"{_fmt(C_mz, 4)}")
    row = _step(ws, row, "·", "k_yy", "", f"{_fmt(k_yy, 4)}")
    row = _step(ws, row, "·", "k_yz", "", f"{_fmt(k_yz, 4)}")
    row = _step(ws, row, "·", "k_zy", "", f"{_fmt(k_zy, 4)}")
    row = _step(ws, row, "·", "k_zz", "", f"{_fmt(k_zz, 4)}")
    row = _blank(ws, row)

    # ── INTERACTION CHECKS ──
    row = _section(ws, row, "INTERACTION CHECKS (EC3 Eq. 6.61 & 6.62)")

    row = _step(ws, row, "Eq. 6.61",
                "N_Ed/N_b,Rd + k_yy × M_y,Ed/M_b,Rd + k_yz × M_z,Ed/M_z,Rd ≤ 1.0",
                f"{_fmt(Ned/1000,2)}/{_fmt(Nbrd/1000,2)} + {_fmt(k_yy)}×{_fmt(Myed/1e6,2)}/{_fmt(Mbrd/1e6,2)} + {_fmt(k_yz)}×{_fmt(Mzed/1e6,2)}/{_fmt(Mzrd/1e6,2)}",
                f"= {_fmt(util_y, 4)}")
    row = _blank(ws, row)

    row = _step(ws, row, "Eq. 6.62",
                "N_Ed/N_b,Rd + k_zy × M_y,Ed/M_b,Rd + k_zz × M_z,Ed/M_z,Rd ≤ 1.0",
                f"{_fmt(Ned/1000,2)}/{_fmt(Nbrd/1000,2)} + {_fmt(k_zy)}×{_fmt(Myed/1e6,2)}/{_fmt(Mbrd/1e6,2)} + {_fmt(k_zz)}×{_fmt(Mzed/1e6,2)}/{_fmt(Mzrd/1e6,2)}",
                f"= {_fmt(util_z, 4)}")
    row = _blank(ws, row)

    row = _blank(ws, row)

    passed = U <= 1.0
    row = _step(ws, row, "Governing",
                "U = max(Eq.6.61, Eq.6.62)",
                f"U = {_fmt(U, 4)}",
                "PASS ✓" if passed else "FAIL ✗")

    row = _blank(ws, row)
    row = _result_row(ws, row,
                      f"Section: {designation}  |  Class {sec_class}  |  Utilisation U = {_fmt(U, 3)}",
                      "PASS ✓" if passed else "FAIL ✗",
                      passed)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# 8. Truss Analysis & Design (batch)
# ─────────────────────────────────────────────

def truss_report(
    member_forces: dict,
    tension_table,          # pandas DataFrame
    compression_table,      # pandas DataFrame
    grade: str,
    shapeten: str,
    shapecomp: str,
    jointing: str,
    nh: float = 0,
    d: float = 0,
    stag: bool = False,
    s: float = 0,
    p: float = 0,
    ngs: float = 0
) -> io.BytesIO:
    wb, ws, row = _new_wb("Truss Analysis & Design — Results Sheet (EC3)")

    fy, fu = get_grade_props(grade)
    E = 210000

    # ── ANALYSIS SUMMARY ──
    row = _section(ws, row, "STRUCTURAL ANALYSIS — Member Axial Forces")
    row = _step(ws, row, "Method", "Direct Stiffness Method (matrix structural analysis)", "", "")
    row = _blank(ws, row)
    row = _step(ws, row, "Member", "Axial Force (kN)", "", "Type")
    for mid, f in member_forces.items():
        mtype = "Tension (+)" if f > 0 else ("Compression (−)" if f < 0 else "Zero-force")
        row = _step(ws, row, f"Member {mid}", f"{_fmt(f, 3)}", "", mtype)

    row = _blank(ws, row)

    # ── DESIGN INPUTS ──
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Steel Grade",     grade)
    row = _step(ws, row, "·", "Tension shape",   shapeten)
    row = _step(ws, row, "·", "Compression shape", shapecomp)
    row = _step(ws, row, "·", "Connection type", jointing)
    if jointing == "bolt":
        row = _step(ws, row, "·", "No. bolt holes n_h",  str(int(nh)))
        row = _step(ws, row, "·", "Bolt diameter d",      f"{_fmt(d)} mm")
        if stag:
            row = _step(ws, row, "·", "Stagger spacing s",  f"{_fmt(s)} mm")
            row = _step(ws, row, "·", "Pitch p",             f"{_fmt(p)} mm")
            row = _step(ws, row, "·", "No. stagger lines",   str(int(ngs)))
    row = _blank(ws, row)

    # ── MATERIAL ──
    row = _section(ws, row, "MATERIAL PROPERTIES")
    row = _step(ws, row, "·", f"f_y ({grade})", "", f"{_fmt(fy)} MPa")
    row = _step(ws, row, "·", f"f_u ({grade})", "", f"{_fmt(fu)} MPa")
    row = _step(ws, row, "·", "E",              "", "210 000 MPa")
    row = _blank(ws, row)

    # ── KEY EQUATIONS ──
    row = _section(ws, row, "KEY DESIGN EQUATIONS")
    row = _step(ws, row, "Tension",
                "N_t,Rd = min(A_gross×f_y/γ_M0 , 0.9×A_net×f_u/γ_M2)",
                "γ_M0=1.0, γ_M2=1.25", "")
    if jointing == "weld":
        row = _step(ws, row, "",
                    "Net area (welded):", "A_net = A_gross", "")
    elif jointing == "bolt" and not stag:
        row = _step(ws, row, "",
                    "Net area (bolted):",
                    "A_net = A_gross − n_h × t × (d+2)", "")
    else:
        row = _step(ws, row, "",
                    "Net area (staggered bolts):",
                    "A_net = A_gross − n_h×t×(d+2) + n_gs×(s²×t)/(4p)", "")
    row = _step(ws, row, "Compression",
                "N_b,Rd = χ × A × f_y / γ_M1    [γ_M1=1.0]", "", "")
    row = _step(ws, row, "",
                "λ̄ = √(A×f_y/N_cr),  N_cr = π²EI/L²", "", "")
    row = _step(ws, row, "",
                "φ = 0.5[1+α(λ̄−0.2)+λ̄²],  χ = 1/(φ+√(φ²−λ̄²))", "", "")
    row = _blank(ws, row)

    # ── TENSION RESULTS ──
    if tension_table is not None and not tension_table.empty:
        row = _section(ws, row, "TENSION MEMBER DESIGN RESULTS")
        # Header
        headers = list(tension_table.columns)
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = _LABEL_FONT
        row += 1
        for _, data_row in tension_table.iterrows():
            for ci, val in enumerate(data_row, start=1):
                ws.cell(row=row, column=ci, value=val).font = _NORMAL_FONT
            row += 1
        row = _blank(ws, row)

    # ── COMPRESSION RESULTS ──
    if compression_table is not None and not compression_table.empty:
        row = _section(ws, row, "COMPRESSION MEMBER DESIGN RESULTS")
        headers = list(compression_table.columns)
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = _LABEL_FONT
        row += 1
        for _, data_row in compression_table.iterrows():
            for ci, val in enumerate(data_row, start=1):
                ws.cell(row=row, column=ci, value=val).font = _NORMAL_FONT
            row += 1

    return _wb_bytes(wb)


# ─────────────────────────────────────────────
# Helper – mirrors truss_analysis.eff_L
# ─────────────────────────────────────────────

def _eff_L(L, endcondition):
    mapping = {
        "Pinned-Pinned": 1.0,
        "Fixed-Fixed":   0.5,
        "Fixed-Pinned":  0.7,
        "Fixed-Free":    2.0,
    }
    factor = mapping.get(endcondition, 1.0)
    return L * factor


# ─────────────────────────────────────────────
# 9. Frame Analysis & Design
# ─────────────────────────────────────────────

def frame_design_report(
    grade: str,
    beam_condition: str,
    col_condition: str,
    col_endcondition: str,
    member_analysis: dict,   # {mid: {"N_start",...,"type","length",...}}
    member_design: dict,     # {mid: design-result dict}
    member_effective_types: dict = None,  # {mid: "Beam"|"Beam-Column"|"Column"}
    nodes=None,
    members=None,
    supports=None,
    node_loads=None,
    udl_loads=None,
    member_point_loads=None,
) -> io.BytesIO:
    """
    Generate a formatted Excel results sheet for the Frame Analysis & Design task.

    member_design values are either the dict returned by restrained_beam (for beams)
    or the dict returned by beam_column (for columns / beam-columns).
    member_effective_types maps each member ID to its actual design category after
    automatic beam-column reclassification.
    """
    wb, ws, row = _new_wb("Frame Analysis & Design — Results Sheet (EC3)")

    fy, fu = get_grade_props(grade)

    # Normalise effective types – fall back to the user-specified analysis type
    # for any member not explicitly recorded (backward-compat with old sessions).
    if member_effective_types is None:
        member_effective_types = {}
    _eff = {
        mid: member_effective_types.get(mid, res["type"])
        for mid, res in member_analysis.items()
    }

    # ── DESIGN INPUTS ─────────────────────────────────────────────────────
    row = _section(ws, row, "DESIGN INPUTS")
    row = _step(ws, row, "·", "Steel Grade",             grade)
    row = _step(ws, row, "·", "f_y",                     f"{_fmt(fy)} MPa")
    row = _step(ws, row, "·", "f_u",                     f"{_fmt(fu)} MPa")
    row = _step(ws, row, "·", "E",                       "210 000 MPa")
    row = _step(ws, row, "·", "Beam Section Type",       "UB  (Universal Beam)")
    row = _step(ws, row, "·", "Beam Lateral Condition",  beam_condition)
    row = _step(ws, row, "·", "Column Section Type",     "UC  (Universal Column)")
    row = _step(ws, row, "·", "Column Steel Condition",  col_condition)
    row = _step(ws, row, "·", "Column End Condition",    col_endcondition)
    row = _blank(ws, row)

    # ── FRAME GEOMETRY ─────────────────────────────────────────────────────
    row = _section(ws, row, "FRAME GEOMETRY — Nodes")
    row = _step(ws, row, "Node", "X (m)", "Y (m)", "")
    for n in nodes:
        row = _step(ws, row, str(int(n[0])), _fmt(float(n[1]), 3), _fmt(float(n[2]), 3), "")
    row = _blank(ws, row)

    row = _section(ws, row, "FRAME GEOMETRY — Members")
    row = _step(ws, row, "Member", "Start Node → End Node", "Type", "Length (m)")
    for m in members:
        mid_label = str(int(m[0]))
        conn      = f"Node {int(m[1])}  →  Node {int(m[2])}"
        mtype     = str(m[3])
        L_m       = member_analysis.get(int(m[0]), {}).get("length", 0.0)
        row = _step(ws, row, mid_label, conn, mtype, _fmt(L_m, 3))
    row = _blank(ws, row)

    row = _section(ws, row, "FRAME GEOMETRY — Support Conditions")
    row = _step(ws, row, "Node", "Condition", "", "")
    for s in supports:
        row = _step(ws, row, str(int(s[0])), str(s[1]), "", "")
    row = _blank(ws, row)

    # ── APPLIED LOADS ─────────────────────────────────────────────────────
    if node_loads:
        row = _section(ws, row, "APPLIED LOADS — Nodal Loads")
        row = _step(ws, row, "Node", "Fx (kN)", "Fy (kN)", "Mz (kNm)")
        for nl in node_loads:
            if any(abs(float(v)) > 1e-12 for v in nl[1:]):
                row = _step(ws, row, str(int(nl[0])),
                            _fmt(float(nl[1]), 3),
                            _fmt(float(nl[2]), 3),
                            _fmt(float(nl[3]), 3))
        row = _blank(ws, row)

    if udl_loads:
        row = _section(ws, row, "APPLIED LOADS — Member UDL")
        row = _step(ws, row, "Member", "wx (kN/m)", "wy (kN/m)", "")
        for ul in udl_loads:
            if any(abs(float(v)) > 1e-12 for v in ul[1:]):
                row = _step(ws, row, str(int(ul[0])),
                            _fmt(float(ul[1]), 3),
                            _fmt(float(ul[2]), 3), "")
        row = _blank(ws, row)

    if member_point_loads:
        row = _section(ws, row, "APPLIED LOADS — Member Point Loads")
        row = _step(ws, row, "Member", "Dist from start (m)", "Fx (kN)", "Fy (kN)")
        for pl in member_point_loads:
            if any(abs(float(v)) > 1e-12 for v in pl[2:]):
                row = _step(ws, row, str(int(pl[0])),
                            _fmt(float(pl[1]), 3),
                            _fmt(float(pl[2]), 3),
                            _fmt(float(pl[3]), 3))
        row = _blank(ws, row)

    # ── ANALYSIS METHOD ────────────────────────────────────────────────────
    row = _section(ws, row, "STRUCTURAL ANALYSIS — Method")
    row = _step(ws, row, "Method",
                "2-D Rigid Frame Direct Stiffness Method",
                "3 DOF per node (u, v, θ)", "")
    row = _step(ws, row, "Beam EI",
                "Assumed representative section for elastic analysis",
                "UB 457×191×67  (I_y = 29 400 cm⁴)", "")
    row = _step(ws, row, "Column EI",
                "Assumed representative section for elastic analysis",
                "UC 254×254×73  (I_y = 11 360 cm⁴)", "")
    row = _blank(ws, row)

    # ── ANALYSIS RESULTS ───────────────────────────────────────────────────
    row = _section(ws, row, "ANALYSIS RESULTS — Member End Forces")
    row = _step(ws, row, "Member", "Type",
                "N_start/N_end (kN)",
                "V_start/V_end (kN)  |  M_start/M_end (kNm)")
    for mid, res in member_analysis.items():
        axial_type = "T (tension)" if res["N_start"] >= 0 else "C (compression)"
        n_str    = f"{_fmt(res['N_start'], 2)} ({axial_type}) / {_fmt(res['N_end'], 2)} kN"
        v_str    = f"{_fmt(res['V_start'], 2)} / {_fmt(res['V_end'], 2)} kN"
        m_str    = f"{_fmt(res['M_start'], 2)} / {_fmt(res['M_end'], 2)} kNm"
        row = _step(ws, row, f"Member {mid}", res["type"], n_str, f"{v_str}  |  {m_str}")
    row = _blank(ws, row)

    # ── DESIGN EQUATIONS ───────────────────────────────────────────────────
    row = _section(ws, row, "DESIGN EQUATIONS")
    row = _step(ws, row, "Beams",
                "Shear:  V_pl,Rd = A_v × f_y / (√3 × γ_M0)   [γ_M0 = 1.0]",
                "", "")
    row = _step(ws, row, "",
                "Moment (restrained): M_pl,Rd = W_pl × f_y / γ_M0",
                "", "")
    row = _step(ws, row, "",
                "Moment (unrestrained): M_b,Rd = χ_LT × W_pl × f_y / γ_M0",
                "", "")
    row = _step(ws, row, "Columns",
                "Axial buckling:  N_b,Rd = χ × A × f_y / γ_M1   [γ_M1 = 1.0]",
                "", "")
    row = _step(ws, row, "",
                "Interaction (EC3 Eq. 6.61): N/N_b,Rd + k_yy × M_y/M_b,Rd + k_yz × M_z/M_z,Rd ≤ 1.0",
                "", "")
    row = _step(ws, row, "Beam-Columns",
                "Members with significant N and M — designed to EC3 §6.3.3.",
                "", "")
    row = _step(ws, row, "",
                "Axial buckling:  N_b,Rd = χ × A × f_y / γ_M1   [γ_M1 = 1.0]",
                "", "")
    row = _step(ws, row, "",
                "LTB resistance:  M_b,Rd = χ_LT × W_pl,y × f_y / γ_M1",
                "", "")
    row = _step(ws, row, "",
                "Interaction (EC3 Eq. 6.61): N/N_b,Rd + k_yy × M_y/M_b,Rd + k_yz × M_z/M_z,Rd ≤ 1.0",
                "", "")
    row = _step(ws, row, "",
                "Interaction (EC3 Eq. 6.62): N/N_b,Rd + k_zy × M_y/M_b,Rd + k_zz × M_z/M_z,Rd ≤ 1.0",
                "", "")
    row = _blank(ws, row)

    # ── DESIGN RESULTS — BEAMS ────────────────────────────────────────────
    beam_mids = [mid for mid in member_design if _eff.get(mid) == "Beam"]
    if beam_mids:
        row = _section(ws, row, "DESIGN RESULTS — Beams (UB sections)")
        row = _step(ws, row, "Member", "Section", "M_Rd (kNm)  /  M_Ed (kNm)",
                    "V_Rd (kN)  /  V_Ed (kN)  |  Utilisation  |  Status")
        for mid in beam_mids:
            dr   = member_design[mid]
            ar   = member_analysis[mid]
            M_Ed = max(abs(ar["M_start"]), abs(ar["M_end"]))
            V_Ed = max(abs(ar["V_start"]), abs(ar["V_end"]))
            Mrd  = dr.get("M_Rd (kNm)", 0.0)
            Vrd  = dr.get("V_Rd (kN)", 0.0)
            util = dr.get("Utilization (%)", 0.0)
            ok   = util <= 100.0
            size = dr.get("Size", "—")
            m_str = f"M_Rd={_fmt(Mrd,2)} / M_Ed={_fmt(M_Ed,2)} kNm"
            v_str = f"V_Rd={_fmt(Vrd,2)} / V_Ed={_fmt(V_Ed,2)} kN | U={_fmt(util,1)}%"
            row = _step(ws, row, f"Member {mid}", size, m_str, v_str)
            passed_str = "PASS ✓" if ok else "FAIL ✗"
            row = _result_row(ws, row,
                              f"Member {mid} — {size}  |  Utilisation {_fmt(util,1)}%",
                              passed_str, ok)
        row = _blank(ws, row)

    # ── DESIGN RESULTS — BEAM-COLUMNS ─────────────────────────────────────
    bc_mids = [mid for mid in member_design if _eff.get(mid) == "Beam-Column"]
    if bc_mids:
        row = _section(ws, row,
                       "DESIGN RESULTS — Beam-Columns (UB sections, EC3 §6.3.3)")
        row = _step(ws, row, "Note",
                    "These members were specified as Beams but carry significant "
                    "axial force alongside bending and have been automatically "
                    "reclassified as Beam-Columns.",
                    "", "")
        row = _step(ws, row, "Member", "Section",
                    "N_b,Rd (kN)  |  N_Ed (kN)  |  M_Ed (kNm)  |  Class",
                    "Utilisation  |  χ_y  |  χ_z  |  χ_LT  |  Status")
        for mid in bc_mids:
            dr    = member_design[mid]
            ar    = member_analysis[mid]
            N_Ed  = max(abs(ar["N_start"]), abs(ar["N_end"]))
            M_Ed  = max(abs(ar["M_start"]), abs(ar["M_end"]))
            U     = dr.get("utilisation", 0.0)
            ok    = U <= 1.0
            size  = dr.get("Designation", "—")
            Nbrd  = dr.get("N_b_Rd", 0.0)
            cls   = dr.get("class", "—")
            chiy  = dr.get("chi_y",  0.0)
            chiz  = dr.get("chi_z",  0.0)
            chiLT = dr.get("chi_LT", 0.0)
            n_str = (f"N_b,Rd={_fmt(Nbrd/1000,2)} kN | N_Ed={_fmt(N_Ed,2)} kN | "
                     f"M_Ed={_fmt(M_Ed,2)} kNm | Class {cls}")
            u_str = (f"U={_fmt(U,3)} | χ_y={_fmt(chiy,3)} | "
                     f"χ_z={_fmt(chiz,3)} | χ_LT={_fmt(chiLT,3)}")
            row = _step(ws, row, f"Member {mid}", size, n_str, u_str)
            passed_str = "PASS ✓" if ok else "FAIL ✗"
            row = _result_row(ws, row,
                              f"Member {mid} — {size}  |  Utilisation {_fmt(U,3)}",
                              passed_str, ok)
        row = _blank(ws, row)

    # ── DESIGN RESULTS — COLUMNS ──────────────────────────────────────────
    col_mids = [mid for mid in member_design if _eff.get(mid) == "Column"]
    if col_mids:
        row = _section(ws, row, "DESIGN RESULTS — Columns (UC sections)")
        row = _step(ws, row, "Member", "Section",
                    "N_b,Rd (kN)  |  N_Ed (kN)  |  Class",
                    "Utilisation  |  χ_y  |  χ_z  |  χ_LT  |  Status")
        for mid in col_mids:
            dr    = member_design[mid]
            ar    = member_analysis[mid]
            N_Ed  = max(abs(ar["N_start"]), abs(ar["N_end"]))
            U     = dr.get("utilisation", 0.0)
            ok    = U <= 1.0
            size  = dr.get("Designation", "—")
            Nbrd  = dr.get("N_b_Rd", 0.0)
            cls   = dr.get("class", "—")
            chiy  = dr.get("chi_y",  0.0)
            chiz  = dr.get("chi_z",  0.0)
            chiLT = dr.get("chi_LT", 0.0)
            n_str = f"N_b,Rd={_fmt(Nbrd/1000,2)} kN | N_Ed={_fmt(N_Ed,2)} kN | Class {cls}"
            u_str = (f"U={_fmt(U,3)} | χ_y={_fmt(chiy,3)} | "
                     f"χ_z={_fmt(chiz,3)} | χ_LT={_fmt(chiLT,3)}")
            row = _step(ws, row, f"Member {mid}", size, n_str, u_str)
            passed_str = "PASS ✓" if ok else "FAIL ✗"
            row = _result_row(ws, row,
                              f"Member {mid} — {size}  |  Utilisation {_fmt(U,3)}",
                              passed_str, ok)
        row = _blank(ws, row)

    # ── DESIGN RESULTS — COLUMNS WITH SIGNIFICANT MOMENTS ─────────────────
    cbc_mids = [mid for mid in member_design if _eff.get(mid) == "Column-BeamColumn"]
    if cbc_mids:
        row = _section(ws, row,
                       "DESIGN RESULTS — Columns with Significant Moments (UC sections, EC3 §6.3.3)")
        row = _step(ws, row, "Note",
                    "These members were specified as Columns but carry significant "
                    "bending moments alongside axial compression. Per EC3 §6.3.3, "
                    "they have been designed as Beam-Columns (N+M interaction check).",
                    "", "")
        row = _step(ws, row, "Member", "Section",
                    "N_b,Rd (kN)  |  N_Ed (kN)  |  M_Ed (kNm)  |  Class",
                    "Utilisation  |  χ_y  |  χ_z  |  χ_LT  |  Status")
        for mid in cbc_mids:
            dr    = member_design[mid]
            ar    = member_analysis[mid]
            N_Ed  = max(abs(ar["N_start"]), abs(ar["N_end"]))
            M_Ed  = max(abs(ar["M_start"]), abs(ar["M_end"]))
            U     = dr.get("utilisation", 0.0)
            ok    = U <= 1.0
            size  = dr.get("Designation", "—")
            Nbrd  = dr.get("N_b_Rd", 0.0)
            cls   = dr.get("class", "—")
            chiy  = dr.get("chi_y",  0.0)
            chiz  = dr.get("chi_z",  0.0)
            chiLT = dr.get("chi_LT", 0.0)
            n_str = (f"N_b,Rd={_fmt(Nbrd/1000,2)} kN | N_Ed={_fmt(N_Ed,2)} kN | "
                     f"M_Ed={_fmt(M_Ed,2)} kNm | Class {cls}")
            u_str = (f"U={_fmt(U,3)} | χ_y={_fmt(chiy,3)} | "
                     f"χ_z={_fmt(chiz,3)} | χ_LT={_fmt(chiLT,3)}")
            row = _step(ws, row, f"Member {mid}", size, n_str, u_str)
            passed_str = "PASS ✓" if ok else "FAIL ✗"
            row = _result_row(ws, row,
                              f"Member {mid} — {size}  |  Utilisation {_fmt(U,3)}",
                              passed_str, ok)
        row = _blank(ws, row)

    return _wb_bytes(wb)
